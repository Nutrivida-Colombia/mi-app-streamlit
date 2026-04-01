
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from datetime import date
import io
import hashlib
import time

try:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

st.set_page_config(
    page_title="NutriVida Colombia",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Paleta nutrición ──────────────────────────────────────────────────────────
# Verde esmeralda (vida, alimentos frescos), Morado (bienestar profesional),
# Dorado (energía, vitalidad), Blanco (limpieza institucional)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@400;600;700&family=DM+Sans:wght@300;400;500&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: "DM Sans", sans-serif; }

.stApp { background: #f8f9f4; color: #1a2010; }

section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1a3a2a 0%, #0f2418 100%);
    border-right: none;
}
section[data-testid="stSidebar"] * { color: #d4edda !important; }
section[data-testid="stSidebar"] .stRadio label {
    background: rgba(255,255,255,0.06); border-radius: 8px;
    padding: 8px 14px !important; font-size: 0.88rem !important;
    transition: background 0.2s; cursor: pointer;
}
section[data-testid="stSidebar"] .stRadio label:hover { background: rgba(255,255,255,0.13); }

.main-header {
    background: linear-gradient(135deg, #1a3a2a 0%, #2d6a4f 40%, #4a9e75 100%);
    border-radius: 16px; padding: 36px 44px; margin-bottom: 28px;
    position: relative; overflow: hidden;
}
.main-header::after {
    content: ""; position: absolute; right: -40px; top: -40px;
    width: 220px; height: 220px;
    background: rgba(255,255,255,0.05); border-radius: 50%;
}
.main-header::before {
    content: ""; position: absolute; right: 80px; bottom: -60px;
    width: 160px; height: 160px;
    background: rgba(255,255,255,0.03); border-radius: 50%;
}
.main-header h1 {
    font-family: "Cormorant Garamond", serif;
    font-size: 2.4rem; font-weight: 700;
    color: #ffffff; margin: 0 0 6px 0; letter-spacing: -0.3px;
}
.main-header p {
    color: rgba(255,255,255,0.65); font-size: 0.82rem;
    margin: 0; font-family: "DM Mono", monospace; letter-spacing: 0.8px;
}

.gov-badge {
    display: inline-block; background: rgba(255,255,255,0.12);
    border: 1px solid rgba(255,255,255,0.2); color: rgba(255,255,255,0.85);
    font-size: 0.68rem; font-family: "DM Mono", monospace;
    padding: 4px 12px; border-radius: 20px; margin-bottom: 12px;
    letter-spacing: 1px; text-transform: uppercase;
}
.author-badge {
    display: inline-block; background: rgba(255,215,0,0.18);
    border: 1px solid rgba(255,215,0,0.35); color: #ffd700;
    font-size: 0.7rem; font-family: "DM Mono", monospace;
    padding: 4px 12px; border-radius: 20px; margin-left: 8px;
}

.logo-mark {
    font-family: "Cormorant Garamond", serif;
    font-size: 1.5rem; font-weight: 700; color: #ffffff;
    line-height: 1.1; letter-spacing: -0.5px;
}
.logo-sub {
    font-size: 0.62rem; color: rgba(255,255,255,0.4);
    font-family: "DM Mono", monospace; letter-spacing: 2px;
    text-transform: uppercase; margin-top: 2px;
}

.section-card {
    background: white; border: 1px solid #e0ead8;
    border-radius: 16px; padding: 24px 28px; margin: 16px 0;
    box-shadow: 0 4px 16px rgba(26,58,42,0.07);
    transition: box-shadow 0.2s;
}
.section-card:hover { box-shadow: 0 6px 24px rgba(26,58,42,0.11); }
.section-card-title {
    font-family: "Cormorant Garamond", serif;
    font-size: 1.1rem; font-weight: 600; color: #1a3a2a;
    margin: 0 0 4px 0; padding-bottom: 10px; border-bottom: 2px solid #e8f5e9;
}
.section-card-subtitle {
    font-size: 0.74rem; color: #5a8a6a;
    font-family: "DM Mono", monospace; margin: 0 0 16px 0;
}
.section-title {
    font-size: 0.68rem; text-transform: uppercase; letter-spacing: 2.5px;
    color: #5a8a6a; font-family: "DM Mono", monospace;
    margin: 28px 0 14px; padding-bottom: 8px; border-bottom: 1px solid #e0ead8;
}

.grupo-etario-badge {
    display: inline-block;
    background: linear-gradient(135deg, #1a3a2a, #2d6a4f);
    color: white; font-size: 0.72rem; font-family: "DM Mono", monospace;
    padding: 4px 14px; border-radius: 20px; margin-bottom: 16px;
}

.result-card { border-radius: 12px; padding: 20px 24px; margin: 12px 0; border-left: 4px solid; }
.result-normal   { background: #f0faf4; border-color: #22c55e; }
.result-normal h3  { color: #15803d; }
.result-normal p   { color: #166534; }
.result-riesgo   { background: #fffbeb; border-color: #f59e0b; }
.result-riesgo h3  { color: #b45309; }
.result-riesgo p   { color: #92400e; }
.result-malo     { background: #fef2f2; border-color: #ef4444; }
.result-malo h3    { color: #b91c1c; }
.result-malo p     { color: #7f1d1d; }
.result-sobrepeso  { background: #faf5ff; border-color: #a855f7; }
.result-sobrepeso h3 { color: #7e22ce; }
.result-sobrepeso p  { color: #6b21a8; }
.result-card h3 { margin: 0 0 6px 0; font-size: 1.05rem; font-family: "Cormorant Garamond", serif; }
.result-card p  { margin: 0; font-size: 0.83rem; }

.recom-item {
    background: white; border: 1px solid #e0ead8; border-radius: 10px;
    padding: 14px 18px; margin: 8px 0; font-size: 0.84rem; color: #2d5a3a;
    box-shadow: 0 2px 8px rgba(26,58,42,0.05);
    transition: box-shadow 0.15s;
}
.recom-item:hover { box-shadow: 0 4px 12px rgba(26,58,42,0.09); }
.recom-item strong { color: #1a3a2a; font-weight: 500; }

.alerta-critica {
    background: #fef2f2; border: 1px solid #fca5a5; border-radius: 10px;
    padding: 14px 18px; font-size: 0.83rem; color: #b91c1c; margin: 8px 0; font-weight: 500;
}
.alerta-moderada {
    background: #fffbeb; border: 1px solid #fcd34d; border-radius: 10px;
    padding: 14px 18px; font-size: 0.83rem; color: #92400e; margin: 8px 0; font-weight: 500;
}

.campo-label {
    font-size: 0.72rem; color: #2d6a4f; font-family: "DM Mono", monospace;
    text-transform: uppercase; letter-spacing: 0.8px; margin-bottom: 2px;
}

.freq-badge { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 0.72rem; font-weight: 500; margin: 2px; font-family: "DM Mono", monospace; }
.freq-diario   { background: #dcfce7; color: #15803d; }
.freq-semanal  { background: #eff6ff; color: #1d4ed8; }
.freq-quincenal{ background: #fef9c3; color: #854d0e; }
.freq-ocasional{ background: #fef2f2; color: #b91c1c; }
.freq-nunca    { background: #f1f5f9; color: #475569; }

.sa-card {
    background: white; border: 1px solid #e0ead8; border-radius: 10px;
    padding: 16px 18px; margin: 8px 0; border-left: 3px solid #2d6a4f;
}
.sa-card-title {
    font-size: 0.78rem; font-weight: 600; color: #1a3a2a;
    font-family: "DM Mono", monospace; text-transform: uppercase;
    letter-spacing: 0.8px; margin-bottom: 10px;
}

.stButton > button {
    background: linear-gradient(135deg, #1a3a2a, #2d6a4f) !important;
    color: white !important; border: none !important; border-radius: 10px !important;
    font-family: "DM Sans", sans-serif !important; font-weight: 500 !important;
    padding: 10px 28px !important; transition: all 0.2s !important;
    box-shadow: 0 4px 12px rgba(26,58,42,0.25) !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 18px rgba(26,58,42,0.35) !important;
}

.stTextInput > div > div > input,
.stNumberInput > div > div > input {
    background: white !important; border: 1.5px solid #c8ddc0 !important;
    border-radius: 10px !important; color: #1a2010 !important;
}
.stSelectbox > div > div {
    background: white !important; border: 1.5px solid #c8ddc0 !important;
    border-radius: 10px !important; color: #1a2010 !important;
}
.stTextArea > div > div > textarea {
    background: white !important; border: 1.5px solid #c8ddc0 !important;
    border-radius: 10px !important; color: #1a2010 !important;
}
.stMultiSelect > div > div {
    background: white !important; border: 1.5px solid #c8ddc0 !important;
    border-radius: 10px !important;
}

div[data-testid="stMetric"] {
    background: white; border: 1px solid #e0ead8; border-radius: 12px;
    padding: 16px 20px; box-shadow: 0 2px 8px rgba(26,58,42,0.06);
}
div[data-testid="stMetric"] label { color: #5a8a6a !important; font-size: 0.78rem !important; }
div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
    color: #1a3a2a !important; font-family: "Cormorant Garamond", serif !important;
}

.stTabs [data-baseweb="tab-list"] {
    background: white; border-radius: 12px; padding: 6px;
    border: 1px solid #e0ead8; gap: 6px;
    box-shadow: 0 2px 8px rgba(26,58,42,0.05);
}
.stTabs [data-baseweb="tab"] {
    color: #5a8a6a !important; border-radius: 8px !important;
    font-family: "DM Sans", sans-serif !important;
    font-size: 0.85rem !important; padding: 8px 16px !important;
    font-weight: 500 !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #1a3a2a, #2d6a4f) !important;
    color: white !important;
    box-shadow: 0 2px 8px rgba(26,58,42,0.25) !important;
}
.stSlider > div > div > div > div { background: #2d6a4f !important; }

.kpi-card {
    background: white; border: 1px solid #e0ead8; border-radius: 14px;
    padding: 20px 16px; text-align: center;
    box-shadow: 0 4px 16px rgba(26,58,42,0.07);
    transition: transform 0.2s, box-shadow 0.2s;
}
.kpi-card:hover { transform: translateY(-2px); box-shadow: 0 8px 24px rgba(26,58,42,0.12); }

.login-container {
    max-width: 420px; margin: 60px auto; padding: 40px;
    background: white; border-radius: 20px;
    border: 1px solid #e0ead8;
    box-shadow: 0 8px 32px rgba(26,58,42,0.10);
}
.login-logo {
    text-align: center; margin-bottom: 28px;
}
.login-title {
    font-family: "Cormorant Garamond", serif;
    font-size: 2rem; font-weight: 700; color: #1a3a2a;
    text-align: center; margin-bottom: 4px;
}
.login-subtitle {
    font-size: 0.78rem; color: #5a8a6a; text-align: center;
    font-family: "DM Mono", monospace; letter-spacing: 0.5px; margin-bottom: 28px;
}

.footer-gov {
    margin-top: 50px; padding: 24px; background: white;
    border: 1px solid #e0ead8; border-radius: 14px;
    font-size: 0.72rem; color: #5a8a6a;
    font-family: "DM Mono", monospace; text-align: center;
    letter-spacing: 0.3px; line-height: 2;
}

@keyframes fadeIn {
    from { opacity: 0; transform: translateY(10px); }
    to   { opacity: 1; transform: translateY(0); }
}
.fade-in { animation: fadeIn 0.5s ease forwards; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SISTEMA DE LOGIN
# ══════════════════════════════════════════════════════════════════════════════
def make_hash(password):
    return hashlib.sha256(str.encode(password)).hexdigest()

USUARIOS = {
    "admin":    make_hash("NutriVida2026"),
    "maira":    make_hash("Nutricionista2026"),
    "demo":     make_hash("Demo2026"),
    "piloto":   make_hash("Piloto2026"),
}

def login():
    st.markdown("""
    <div style="display:flex;justify-content:center;align-items:center;min-height:80vh;">
    <div class="login-container fade-in">
        <div class="login-logo">
            <div style="width:72px;height:72px;background:linear-gradient(135deg,#1a3a2a,#2d6a4f);
                        border-radius:50%;margin:0 auto 12px;display:flex;align-items:center;
                        justify-content:center;">
                <span style="font-size:32px;">🌿</span>
            </div>
            <div class="login-title">NutriVida Colombia</div>
            <div class="login-subtitle">Sistema Integral de Evaluación y Seguimiento Nutricional</div>
            <div style="display:inline-block;background:#e8f5e9;border:1px solid #c8ddc0;
                        color:#1a3a2a;font-size:0.65rem;font-family:'DM Mono',monospace;
                        padding:3px 10px;border-radius:20px;letter-spacing:1px;">
                v7.0 · MinSalud / ICBF · 2026
            </div>
        </div>
    </div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown('<div class="campo-label">Usuario</div>', unsafe_allow_html=True)
        username = st.text_input("", placeholder="Ingresa tu usuario", key="usr", label_visibility="collapsed")
        st.markdown('<div class="campo-label" style="margin-top:10px;">Contraseña</div>', unsafe_allow_html=True)
        password = st.text_input("", placeholder="Ingresa tu contraseña", type="password", key="pwd", label_visibility="collapsed")
        st.markdown("<br>", unsafe_allow_html=True)

        if st.button("Ingresar al sistema", key="login_btn"):
            if username in USUARIOS and USUARIOS[username] == make_hash(password):
                st.session_state["logged_in"] = True
                st.session_state["username"] = username
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos")

        st.markdown("""
        <div style="text-align:center;margin-top:20px;font-size:0.7rem;color:#5a8a6a;
                    font-family:'DM Mono',monospace;line-height:1.8;">
            © 2026 Maira Alejandra Carrillo Florez<br>
            Acceso restringido a personal autorizado<br>
            Ley 23 de 1982 · Decisión Andina 351
        </div>
        """, unsafe_allow_html=True)

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login()
    st.stop()


# ── Grupos etários Colombia — Normativa vigente ────────────────────────────────
# Ley 1098/2006 (Código Infancia), Ley 1622/2013 (Estatuto Juventud),
# CONPES 3873/2016, Política Nacional Envejecimiento 2007-2019
GRUPOS_ETARIOS = {
    "Primera infancia (0-5 años)": {
        "rango": (0, 5),
        "norma": "Ley 1098/2006 — Código de Infancia y Adolescencia",
        "indicadores": ["Peso/Edad","Talla/Edad","IMC/Edad","Perímetro cefálico (< 2 años)","Perímetro braquial"],
        "color": "#22c55e",
    },
    "Infancia (6-11 años)": {
        "rango": (6, 11),
        "norma": "Ley 1098/2006 — Código de Infancia y Adolescencia",
        "indicadores": ["IMC/Edad","Talla/Edad","Perímetro cintura"],
        "color": "#3b82f6",
    },
    "Adolescencia (12-17 años)": {
        "rango": (12, 17),
        "norma": "Ley 1098/2006 — Código de Infancia y Adolescencia",
        "indicadores": ["IMC/Edad","Talla/Edad","Perímetro cintura","Maduración sexual (Tanner)"],
        "color": "#8b5cf6",
    },
    "Juventud (18-28 años)": {
        "rango": (18, 28),
        "norma": "Ley 1622/2013 — Estatuto de Ciudadanía Juvenil",
        "indicadores": ["IMC","Perímetro cintura","Índice cintura/cadera","GET"],
        "color": "#f59e0b",
    },
    "Adultez (29-59 años)": {
        "rango": (29, 59),
        "norma": "Ley 1751/2015 — Ley Estatutaria de Salud",
        "indicadores": ["IMC","Perímetro cintura","Índice cintura/cadera","GET","% grasa corporal"],
        "color": "#f97316",
    },
    "Persona mayor (≥60 años)": {
        "rango": (60, 120),
        "norma": "Ley 1251/2008 — Política Nacional de Envejecimiento",
        "indicadores": ["IMC","MNA (Mini Nutritional Assessment)","Perímetro pantorrilla","Fuerza de agarre"],
        "color": "#ef4444",
    },
}

FRECUENCIAS_CONSUMO = [
    "Diario","5-6 veces/semana","3-4 veces/semana",
    "1-2 veces/semana","Quincenal","Mensual","Ocasional","Nunca"
]

GRUPOS_ALIMENTOS = {
    "Cereales, raíces, tubérculos y plátanos": ["Arroz","Pan","Papa","Yuca","Plátano","Pasta","Avena"],
    "Hortalizas y verduras": ["Zanahoria","Espinaca","Tomate","Cebolla","Brócoli","Pepino","Lechuga"],
    "Frutas": ["Banano","Mango","Naranja","Guayaba","Papaya","Mora","Maracuyá"],
    "Carnes, huevo, leguminosas y mezclas vegetales": ["Res","Cerdo","Pollo","Pescado","Huevo","Fríjol","Lenteja","Garbanzo"],
    "Lácteos y derivados": ["Leche","Yogur","Queso","Kumis","Suero"],
    "Grasas": ["Aceite vegetal","Mantequilla","Margarina","Aguacate","Maní"],
    "Azúcares y dulces": ["Azúcar","Panela","Miel","Mermelada","Gaseosa","Jugo artificial"],
    "Comidas rápidas y ultraprocesados": ["Hamburguesa","Pizza","Papas fritas","Embutidos","Snacks","Bebidas energizantes"],
}

DEPARTAMENTOS = {
    "Amazonas":["Leticia","Puerto Nariño","El Encanto"],
    "Antioquia":["Medellín","Bello","Itagüí","Envigado","Apartadó","Turbo","Rionegro","Caucasia"],
    "Arauca":["Arauca","Saravena","Tame","Fortul"],
    "Atlántico":["Barranquilla","Soledad","Malambo","Sabanalarga"],
    "Bolívar":["Cartagena","Magangué","El Carmen de Bolívar","Mompox"],
    "Boyacá":["Tunja","Duitama","Sogamoso","Chiquinquirá"],
    "Caldas":["Manizales","Villamaría","La Dorada","Chinchiná"],
    "Caquetá":["Florencia","San Vicente del Caguán","Puerto Rico"],
    "Casanare":["Yopal","Aguazul","Villanueva","Tauramena"],
    "Cauca":["Popayán","Santander de Quilichao","Puerto Tejada","Patía"],
    "Cesar":["Valledupar","Aguachica","Bosconia","Codazzi"],
    "Chocó":["Quibdó","Istmina","Tadó","Condoto","Acandí"],
    "Córdoba":["Montería","Cereté","Lorica","Tierralta","Sahagún"],
    "Cundinamarca":["Bogotá D.C.","Soacha","Facatativá","Zipaquirá","Chía"],
    "Guainía":["Inírida","Barranco Minas"],
    "Guaviare":["San José del Guaviare","Calamar","El Retorno"],
    "Huila":["Neiva","Pitalito","Garzón","La Plata"],
    "La Guajira":["Riohacha","Maicao","Uribia","Manaure","Fonseca"],
    "Magdalena":["Santa Marta","Ciénaga","Fundación","El Banco"],
    "Meta":["Villavicencio","Acacías","Granada","Puerto López"],
    "Nariño":["Pasto","Tumaco","Ipiales","Barbacoas","Túquerres"],
    "Norte de Santander":["Cúcuta","Ocaña","Pamplona","Villa del Rosario"],
    "Putumayo":["Mocoa","Puerto Asís","Orito","Valle del Guamuez"],
    "Quindío":["Armenia","Calarcá","Montenegro","Quimbaya"],
    "Risaralda":["Pereira","Dosquebradas","Santa Rosa de Cabal"],
    "San Andrés y Providencia":["San Andrés","Providencia"],
    "Santander":["Bucaramanga","Floridablanca","Girón","Piedecuesta","Barrancabermeja"],
    "Sucre":["Sincelejo","Corozal","Sampués","San Marcos"],
    "Tolima":["Ibagué","Espinal","Melgar","Honda"],
    "Valle del Cauca":["Cali","Buenaventura","Palmira","Tuluá","Buga"],
    "Vaupés":["Mitú","Carurú"],
    "Vichada":["Puerto Carreño","La Primavera","Santa Rosalía"],
}

# ── Funciones clínicas ─────────────────────────────────────────────────────────
def calcular_imc(peso, talla_cm):
    return round(peso / (talla_cm/100)**2, 1)

def determinar_grupo_etario(edad_anos):
    for grupo, data in GRUPOS_ETARIOS.items():
        a, b = data["rango"]
        if a <= edad_anos <= b:
            return grupo, data
    return "Adultez (29-59 años)", GRUPOS_ETARIOS["Adultez (29-59 años)"]

def clasificar_adulto(imc):
    if imc < 16:     return ("Desnutrición severa","malo","Requiere atención urgente — riesgo vital")
    elif imc < 17:   return ("Desnutrición moderada","malo","Intervención nutricional prioritaria")
    elif imc < 18.5: return ("Desnutrición leve","riesgo","Seguimiento y plan de recuperación")
    elif imc < 25:   return ("Estado nutricional normal","normal","Mantener hábitos actuales")
    elif imc < 30:   return ("Sobrepeso","sobrepeso","Orientación alimentaria y actividad física")
    elif imc < 35:   return ("Obesidad grado I","riesgo","Plan nutricional y seguimiento médico")
    elif imc < 40:   return ("Obesidad grado II","malo","Intervención multidisciplinaria")
    else:            return ("Obesidad mórbida grado III","malo","Atención especializada urgente")

def clasificar_nino(imc, edad_meses):
    if edad_meses < 24:    lim=(14.0,18.0)
    elif edad_meses < 60:  lim=(13.5,17.5)
    elif edad_meses < 120: lim=(13.0,19.0)
    else:                  lim=(14.0,22.0)
    if imc < lim[0]-2:    return ("Desnutrición aguda severa","malo","Remisión urgente — protocolo AIEPI")
    elif imc < lim[0]:    return ("Desnutrición aguda moderada","riesgo","Seguimiento semanal + suplementación")
    elif imc <= lim[1]:   return ("Eutrófico (normal)","normal","Continuar monitoreo de rutina")
    elif imc <= lim[1]+2: return ("Riesgo de sobrepeso","riesgo","Orientación alimentaria familiar")
    else:                 return ("Obesidad infantil","sobrepeso","Evaluación pediátrica y nutricional")

def peso_ideal(talla, sexo):
    if sexo=="Masculino": return round(talla-100-(talla-150)/4,1)
    return round(talla-100-(talla-150)/2,1)

def get_energia(peso, talla, edad, sexo, actividad):
    tmb=(10*peso+6.25*talla-5*edad+5) if sexo=="Masculino" else (10*peso+6.25*talla-5*edad-161)
    f={"Sedentario":1.2,"Ligero":1.375,"Moderado":1.55,"Intenso":1.725,"Muy intenso":1.9}
    return round(tmb*f.get(actividad,1.55))

def recomendaciones(clf):
    r={
        "Desnutrición severa":[("Remisión inmediata","Referir a hospital de referencia para manejo de desnutrición aguda severa"),("Protocolo AIEPI","Aplicar protocolo de atención integrada"),("ATRC","Alimento terapéutico listo para consumo según peso"),("Seguimiento 48h","Visita domiciliaria en menos de 48 horas"),("SIVIGILA","Notificación obligatoria inmediata")],
        "Desnutrición moderada":[("Suplementación","Vitamina A, hierro, zinc y ácido fólico según protocolo MinSalud"),("Plan alimentario","Dieta hipercalórica e hiperproteica — 5 comidas/día"),("Seguimiento semanal","Control de peso y talla cada 7 días"),("SISVAN","Registrar en sistema de vigilancia nutricional")],
        "Desnutrición leve":[("Orientación nutricional","Consejería familiar sobre grupos alimentarios y porciones"),("Seguimiento mensual","Control nutricional cada 30 días"),("Programas ICBF","Vincular a programas de complementación alimentaria")],
        "Estado nutricional normal":[("Mantener hábitos","Reforzar alimentación saludable y actividad física"),("Control periódico","Evaluación nutricional cada 6 meses")],
        "Sobrepeso":[("Plan alimentario","Reducción calórica gradual — déficit 300-500 kcal/día"),("Actividad física","Mínimo 150 min/semana de actividad moderada"),("Seguimiento mensual","Control mensual de peso y medidas")],
        "Obesidad grado I":[("Multidisciplinario","Nutricionista + médico + actividad física"),("Laboratorios","Perfil lipídico, glucemia, HbA1c, TSH"),("Seguimiento quincenal","Control cada 15 días primeros 3 meses")],
        "Obesidad grado II":[("Evaluación especializada","Remisión a nutricionista clínico y medicina interna"),("Laboratorios completos","Perfil lipídico, glucemia, función hepática y renal")],
        "Obesidad mórbida grado III":[("Remisión urgente","Evaluación por cirugía bariátrica y endocrinología")],
        "Desnutrición aguda severa":[("Remisión urgente","Hospital de referencia — protocolo OMS fase 1 y 2"),("F-75 y F-100","Fórmula terapéutica según peso y estado clínico"),("SIVIGILA","Notificación obligatoria"),("Cuidador","Capacitación al cuidador en preparación de alimentos")],
        "Desnutrición aguda moderada":[("ATRC domiciliario","Alimento terapéutico listo para consumo — dosis según peso"),("Control semanal","Peso y talla cada semana — meta 5-10g/kg/día"),("Tamizaje familiar","Evaluar estado nutricional de otros niños en el hogar")],
        "Eutrófico (normal)":[("Control rutinario","Siguiente control según esquema de crecimiento y desarrollo"),("Alimentación","Reforzar prácticas de alimentación complementaria adecuada")],
        "Riesgo de sobrepeso":[("Orientación familiar","Consejería sobre porciones y alimentos ultraprocesados"),("Actividad física","60 minutos diarios de actividad física para niños")],
        "Obesidad infantil":[("Evaluación pediátrica","Descartar causas endocrinas — TSH, cortisol"),("Plan familiar","Modificación de hábitos en todo el núcleo familiar")],
    }
    for k in r:
        if k in clf: return r[k]
    return [("Evaluación adicional","Consultar con especialista en nutrición clínica")]

@st.cache_data
def datos_demo():
    np.random.seed(42); n=320
    deptos=list(DEPARTAMENTOS.keys())
    ds=np.random.choice(deptos,n)
    ms=[np.random.choice(DEPARTAMENTOS[d]) for d in ds]
    sx=np.random.choice(["Masculino","Femenino"],n)
    ed=np.random.randint(6,60,n)
    pw=np.where(sx=="Masculino",np.random.normal(14.2,2.8,n),np.random.normal(13.6,2.5,n))
    tl=np.random.normal(93,19,n)
    iv=pw/((tl/100)**2)
    est=[]
    for v in iv:
        if v<13: est.append("Desnutrición severa")
        elif v<15: est.append("Desnutrición moderada")
        elif v<17: est.append("Desnutrición leve")
        elif v<18.5: est.append("Normal")
        elif v<20: est.append("Riesgo sobrepeso")
        elif v<25: est.append("Sobrepeso")
        else: est.append("Obesidad")
    fechas=pd.date_range("2024-01-01","2025-03-01",periods=n)
    grupos_e=[]
    for e in ed:
        ea=e//12
        g,_=determinar_grupo_etario(ea)
        grupos_e.append(g)
    return pd.DataFrame({"fecha":fechas,"departamento":ds,"municipio":ms,"edad_meses":ed,"sexo":sx,"peso_kg":np.round(pw,1),"talla_cm":np.round(tl,1),"imc":np.round(iv,1),"estado_nutricional":est,"grupo_etario":grupos_e})

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""<div style="padding:20px 0 16px;">
    <div style="font-size:0.62rem;color:rgba(255,255,255,0.35);font-family:'DM Mono',monospace;letter-spacing:2px;text-transform:uppercase;margin-bottom:10px;">Sistema Nacional</div>
    <div style="font-family:'Playfair Display',serif;font-size:1.3rem;font-weight:700;color:white;line-height:1.2;">NutriVida<br>Colombia</div>
    <div style="font-size:0.68rem;color:rgba(255,255,255,0.4);font-family:'DM Mono',monospace;letter-spacing:1px;margin-top:4px;">v7.0 · MinSalud / ICBF · 2026</div>
    </div><hr style="border-color:rgba(255,255,255,0.1);margin:0 0 16px 0;">""", unsafe_allow_html=True)
    modulo = st.radio("Módulo",["Evaluación Integral","Referencia Epidemiológica","Dashboard Poblacional","Registro Masivo","Acerca del sistema"],label_visibility="collapsed")
    st.markdown("""<hr style="border-color:rgba(255,255,255,0.1);margin:20px 0 12px;">
    <div style="font-size:0.65rem;color:rgba(255,255,255,0.3);font-family:'DM Mono',monospace;line-height:2;letter-spacing:0.5px;">
    Ley 1098/2006 · Ley 1622/2013<br>Ley 1251/2008 · OMS/OPS 2006<br>ENSIN 2015 · Res. 2465/2016<br>ELCSA · SISVAN · SIVIGILA
    </div>""", unsafe_allow_html=True)

# ── Header ─────────────────────────────────────────────────────────────────────
# Show logged in user
    username = st.session_state.get("username","")
    if st.sidebar.button("Cerrar sesión"):
        st.session_state["logged_in"] = False
        st.rerun()
    st.sidebar.markdown(f"""<div style="font-size:0.65rem;color:rgba(255,255,255,0.4);font-family:'DM Mono',monospace;margin-top:8px;">
    Sesión: {username}</div>""", unsafe_allow_html=True)

st.markdown("""<div class="main-header fade-in">
<div><span class="gov-badge">República de Colombia · Ministerio de Salud · ICBF</span>
<span class="author-badge">© Maira Alejandra Carrillo Florez</span></div>
<h1>Sistema Integral de Evaluación y Seguimiento Nutricional</h1>
<p>NUTRIVIDA v7.0 · Grupos etarios Normativa Colombia · 32 departamentos · 2026</p>
</div>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 1 — EVALUACIÓN INTEGRAL
# ═══════════════════════════════════════════════════════════════════════════════
if modulo == "Evaluación Integral":

    st.markdown('<div class="section-title">Selección de grupo etario — Normativa Colombia vigente</div>', unsafe_allow_html=True)
    st.markdown('<div class="recom-item"><strong>Clasificación según:</strong> Ley 1098/2006 (Código Infancia y Adolescencia), Ley 1622/2013 (Estatuto Juventud), Ley 1251/2008 (Envejecimiento), Ley 1751/2015 (Ley Estatutaria de Salud)</div>', unsafe_allow_html=True)

    grupo_sel = st.selectbox("Seleccionar grupo etario del paciente", list(GRUPOS_ETARIOS.keys()))
    info_grupo = GRUPOS_ETARIOS[grupo_sel]

    st.markdown(f"""
    <div style="background:white;border:1px solid #e8e0f5;border-radius:10px;padding:12px 18px;margin:8px 0;display:flex;align-items:center;gap:12px;">
        <div style="background:{info_grupo['color']};width:10px;height:10px;border-radius:50%;flex-shrink:0;"></div>
        <div>
            <div style="font-size:0.75rem;font-weight:500;color:#2d1b69;">{grupo_sel}</div>
            <div style="font-size:0.7rem;color:#9580c0;font-family:'DM Mono',monospace;">{info_grupo['norma']}</div>
        </div>
    </div>""", unsafe_allow_html=True)

    es_menor = info_grupo["rango"][1] <= 17

    # ── 1. IDENTIFICACIÓN ─────────────────────────────────────────────────────
    st.markdown("""<div class="section-card">
    <div class="section-card-title">1. Datos de identificación</div>
    <div class="section-card-subtitle">Información básica del paciente para el registro y seguimiento</div>""", unsafe_allow_html=True)
    c1,c2,c3 = st.columns(3)
    with c1:
        st.markdown('<div class="campo-label">Nombre completo</div>', unsafe_allow_html=True)
        nombre = st.text_input("", key="nom", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Número de documento</div>', unsafe_allow_html=True)
        doc = st.text_input("", key="doc", label_visibility="collapsed")
    with c2:
        st.markdown('<div class="campo-label">Fecha de evaluación</div>', unsafe_allow_html=True)
        fecha_ev = st.date_input("", value=date.today(), key="fec", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Sexo biológico</div>', unsafe_allow_html=True)
        sexo = st.selectbox("", ["Femenino","Masculino"], key="sex", label_visibility="collapsed")
    with c3:
        st.markdown('<div class="campo-label">Departamento</div>', unsafe_allow_html=True)
        depto = st.selectbox("", sorted(DEPARTAMENTOS.keys()), key="dep", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Municipio</div>', unsafe_allow_html=True)
        mun = st.selectbox("", DEPARTAMENTOS[depto], key="mun", label_visibility="collapsed")
    if es_menor:
        c1,c2 = st.columns(2)
        with c1:
            st.markdown('<div class="campo-label">Nombre del cuidador o acudiente</div>', unsafe_allow_html=True)
            cuidador = st.text_input("", key="cui", label_visibility="collapsed")
        with c2:
            st.markdown('<div class="campo-label">Parentesco con el menor</div>', unsafe_allow_html=True)
            parentesco = st.selectbox("", ["Madre","Padre","Abuelo/a","Tío/a","Hermano/a mayor","Otro"], key="par", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

    # ── 2. DATOS ANTROPOMÉTRICOS ──────────────────────────────────────────────
    st.markdown(f"""<div class="section-card">
    <div class="section-card-title">2. Datos antropométricos</div>
    <div class="section-card-subtitle">Indicadores para este grupo: {' · '.join(info_grupo['indicadores'])}</div>""", unsafe_allow_html=True)

    c1,c2,c3,c4 = st.columns(4)
    with c1:
        st.markdown('<div class="campo-label">Edad (años)</div>', unsafe_allow_html=True)
        edad_a = st.number_input("", 0, 120, 25 if not es_menor else 4, key="eda", label_visibility="collapsed")
        if es_menor:
            st.markdown('<div class="campo-label">Meses adicionales</div>', unsafe_allow_html=True)
            edad_m_ad = st.number_input("", 0, 11, 0, key="ema", label_visibility="collapsed")
    with c2:
        st.markdown('<div class="campo-label">Peso (kg)</div>', unsafe_allow_html=True)
        peso = st.number_input("", 2.0, 300.0, 65.0 if not es_menor else 16.5, step=0.1, key="pes", label_visibility="collapsed")
    with c3:
        st.markdown('<div class="campo-label">Talla (cm)</div>', unsafe_allow_html=True)
        talla = st.number_input("", 40.0, 220.0, 162.0 if not es_menor else 103.0, step=0.5, key="tal", label_visibility="collapsed")
    with c4:
        if not es_menor:
            st.markdown('<div class="campo-label">Nivel de actividad física</div>', unsafe_allow_html=True)
            actividad = st.selectbox("", ["Sedentario","Ligero","Moderado","Intenso","Muy intenso"], key="act", label_visibility="collapsed")

    st.markdown('<div style="margin-top:12px;"></div>', unsafe_allow_html=True)
    c1,c2,c3,c4 = st.columns(4)
    with c1:
        st.markdown('<div class="campo-label">Perímetro braquial (cm) — opcional</div>', unsafe_allow_html=True)
        p_brazo = st.number_input("", 0.0, 50.0, 0.0, step=0.1, key="pba", label_visibility="collapsed", help="Útil en < 5 años y adultos mayores")
    with c2:
        st.markdown('<div class="campo-label">Perímetro cintura (cm) — opcional</div>', unsafe_allow_html=True)
        p_cintura = st.number_input("", 0.0, 200.0, 0.0, step=0.5, key="pci", label_visibility="collapsed", help="Riesgo cardiovascular y metabólico")
    with c3:
        st.markdown('<div class="campo-label">Perímetro cadera (cm) — opcional</div>', unsafe_allow_html=True)
        p_cadera = st.number_input("", 0.0, 200.0, 0.0, step=0.5, key="pca", label_visibility="collapsed", help="Para cálculo índice cintura/cadera")
    with c4:
        if grupo_sel == "Persona mayor (≥60 años)":
            st.markdown('<div class="campo-label">Perímetro pantorrilla (cm) — opcional</div>', unsafe_allow_html=True)
            p_pantorrilla = st.number_input("", 0.0, 60.0, 0.0, step=0.5, key="ppa", label_visibility="collapsed", help="Indicador sarcopenia en adulto mayor")
        elif es_menor and info_grupo["rango"][0] == 0:
            st.markdown('<div class="campo-label">Perímetro cefálico (cm) — opcional</div>', unsafe_allow_html=True)
            p_cefalico = st.number_input("", 0.0, 60.0, 0.0, step=0.5, key="pce", label_visibility="collapsed", help="Obligatorio en < 2 años")
    st.markdown("</div>", unsafe_allow_html=True)

    # ── 3. DATOS SOCIODEMOGRÁFICOS ────────────────────────────────────────────
    st.markdown("""<div class="section-card">
    <div class="section-card-title">3. Datos sociodemográficos</div>
    <div class="section-card-subtitle">Determinantes sociales relacionados con el estado nutricional</div>""", unsafe_allow_html=True)
    c1,c2,c3 = st.columns(3)
    with c1:
        st.markdown('<div class="campo-label">Estrato socioeconómico</div>', unsafe_allow_html=True)
        estrato = st.selectbox("", ["1 — Bajo-bajo","2 — Bajo","3 — Medio-bajo","4 — Medio","5 — Medio-alto","6 — Alto","Sin estrato"], key="est", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Régimen de salud</div>', unsafe_allow_html=True)
        regimen = st.selectbox("", ["Subsidiado (SISBEN)","Contributivo","Especial (Fuerzas militares, Ecopetrol)","No asegurado","Desconocido"], key="reg", label_visibility="collapsed")
    with c2:
        st.markdown('<div class="campo-label">Tipo de vivienda</div>', unsafe_allow_html=True)
        vivienda = st.selectbox("", ["Casa propia","Arrendada","Familiar (sin pago)","Invasión o asentamiento informal","Otro"], key="viv", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Número de personas en el hogar</div>', unsafe_allow_html=True)
        personas = st.number_input("", 1, 20, 4, key="per", label_visibility="collapsed")
    with c3:
        st.markdown('<div class="campo-label">Nivel educativo del cuidador o paciente</div>', unsafe_allow_html=True)
        escolaridad = st.selectbox("", ["Sin escolaridad","Primaria incompleta","Primaria completa","Secundaria incompleta","Secundaria completa","Técnico o tecnólogo","Universitario","Posgrado"], key="esc", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Pertenencia étnica</div>', unsafe_allow_html=True)
        etnia = st.selectbox("", ["Ninguna","Indígena","Afrocolombiano","Raizal","Palenquero","Gitano o Rom","Otro"], key="etn", label_visibility="collapsed")
    st.markdown('<div class="campo-label" style="margin-top:12px;">Número de comidas al día</div>', unsafe_allow_html=True)
    comidas = st.slider("", 1, 8, 3, key="cmd", label_visibility="collapsed", help="Incluye todas las comidas principales y refrigerios")
    st.caption(f"El paciente consume {comidas} comida(s) al día — {'Adecuado' if comidas >= 3 else 'Por debajo de lo recomendado (mínimo 3)'}")
    st.markdown("</div>", unsafe_allow_html=True)

    # ── 4. SEGURIDAD ALIMENTARIA — 5 COMPONENTES ──────────────────────────────
    st.markdown("""<div class="section-card">
    <div class="section-card-title">4. Seguridad alimentaria y nutricional</div>
    <div class="section-card-subtitle">Evaluación de los 5 componentes — Política Nacional de Seguridad Alimentaria y Nutricional (CONPES 113/2007)</div>""", unsafe_allow_html=True)

    # Componente 1: Disponibilidad
    st.markdown('<div class="sa-card"><div class="sa-card-title">Componente 1 — Disponibilidad de alimentos</div>', unsafe_allow_html=True)
    st.caption("Suficiencia de alimentos en el hogar para cubrir las necesidades de todos los miembros")
    c1,c2 = st.columns(2)
    with c1:
        st.markdown('<div class="campo-label">Disponibilidad de alimentos en el hogar</div>', unsafe_allow_html=True)
        disponibilidad = st.selectbox("", ["Siempre hay alimentos suficientes","Casi siempre hay suficientes","A veces hay suficientes","Casi nunca hay suficientes","Nunca hay alimentos suficientes"], key="dis", label_visibility="collapsed")
    with c2:
        st.markdown('<div class="campo-label">Principal fuente de obtención de alimentos</div>', unsafe_allow_html=True)
        fuente_alimentos = st.selectbox("", ["Compra en tienda o supermercado","Producción propia (huerta, finca)","Donación o ayuda humanitaria","Programas sociales (ICBF, PAE)","Mixta (compra + producción)"], key="fue", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

    # Componente 2: Acceso
    st.markdown('<div class="sa-card"><div class="sa-card-title">Componente 2 — Acceso económico y físico</div>', unsafe_allow_html=True)
    st.caption("Capacidad del hogar para obtener alimentos suficientes y nutritivos")
    c1,c2 = st.columns(2)
    with c1:
        st.markdown('<div class="campo-label">¿El ingreso familiar alcanza para alimentación?</div>', unsafe_allow_html=True)
        acceso_eco = st.selectbox("", ["Siempre alcanza","Casi siempre alcanza","A veces no alcanza","Frecuentemente no alcanza","Nunca alcanza"], key="ace", label_visibility="collapsed")
    with c2:
        st.markdown('<div class="campo-label">Distancia al punto de venta de alimentos más cercano</div>', unsafe_allow_html=True)
        acceso_fisico = st.selectbox("", ["Menos de 15 minutos","15-30 minutos","30-60 minutos","Más de 1 hora","Sin acceso (zona rural remota)"], key="acf", label_visibility="collapsed")
    prog_alim = st.multiselect("Programas alimentarios activos en el hogar", ["PAE — Programa de Alimentación Escolar","Comedores comunitarios ICBF","Familias en Acción — componente nutrición","Subsidio de desayunos infantiles","Banco de alimentos","Ninguno"], key="prog")
    st.markdown("</div>", unsafe_allow_html=True)

    # Componente 3: Consumo
    st.markdown('<div class="sa-card"><div class="sa-card-title">Componente 3 — Consumo y hábitos alimentarios</div>', unsafe_allow_html=True)
    st.caption("Patrones de alimentación, selección y preparación de alimentos en el hogar")
    c1,c2 = st.columns(2)
    with c1:
        st.markdown('<div class="campo-label">¿Quién toma decisiones sobre la alimentación en el hogar?</div>', unsafe_allow_html=True)
        decision_alim = st.selectbox("", ["La madre","El padre","Ambos padres","El cuidador principal","El paciente mismo","Otro"], key="dec", label_visibility="collapsed")
    with c2:
        st.markdown('<div class="campo-label">Nivel de inseguridad alimentaria (ELCSA)</div>', unsafe_allow_html=True)
        inseguridad = st.selectbox("", ["Seguridad alimentaria","Inseguridad leve","Inseguridad moderada","Inseguridad severa","No evaluada"], key="ins", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

    # Componente 4: Calidad e inocuidad
    st.markdown('<div class="sa-card"><div class="sa-card-title">Componente 4 — Calidad e inocuidad de los alimentos</div>', unsafe_allow_html=True)
    st.caption("Condiciones higiénicas y sanitarias en la preparación y conservación de alimentos")
    c1,c2,c3 = st.columns(3)
    with c1:
        st.markdown('<div class="campo-label">Acceso a agua potable</div>', unsafe_allow_html=True)
        agua = st.selectbox("", ["Acueducto certificado","Agua comprada en botellón","Agua de pozo con tratamiento","Agua de río, lluvia o sin tratamiento","Sin acceso a agua segura"], key="agu", label_visibility="collapsed")
    with c2:
        st.markdown('<div class="campo-label">Condiciones de almacenamiento de alimentos</div>', unsafe_allow_html=True)
        almacenamiento = st.selectbox("", ["Adecuado (nevera + despensa)","Parcial (solo nevera o solo despensa)","Inadecuado (sin refrigeración)","Sin condiciones mínimas"], key="alm", label_visibility="collapsed")
    with c3:
        st.markdown('<div class="campo-label">Condiciones de preparación de alimentos</div>', unsafe_allow_html=True)
        preparacion = st.selectbox("", ["Cocina con todos los servicios","Cocina básica (estufa + agua)","Fogón de leña o carbón","Sin condiciones adecuadas"], key="pre", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

    # Componente 5: Aprovechamiento biológico
    st.markdown('<div class="sa-card"><div class="sa-card-title">Componente 5 — Aprovechamiento biológico (utilización)</div>', unsafe_allow_html=True)
    st.caption("Capacidad del organismo para absorber y utilizar los nutrientes de los alimentos")
    c1,c2 = st.columns(2)
    with c1:
        st.markdown('<div class="campo-label">Estado de salud general que afecta absorción de nutrientes</div>', unsafe_allow_html=True)
        salud_absorcion = st.selectbox("", ["Sin condiciones que afecten absorción","Infecciones recurrentes (diarrea, respiratorias)","Parasitosis intestinal","Enfermedad crónica que afecta absorción","Múltiples condiciones"], key="sal", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Última desparasitación</div>', unsafe_allow_html=True)
        despar = st.selectbox("", ["Hace menos de 6 meses","Hace 6-12 meses","Hace más de 1 año","No recuerda","Nunca"], key="des", label_visibility="collapsed")
    with c2:
        st.markdown('<div class="campo-label">Acceso a saneamiento básico</div>', unsafe_allow_html=True)
        saneamiento = st.selectbox("", ["Alcantarillado y baño privado","Pozo séptico","Letrina","Sin saneamiento básico"], key="san", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Lavado de manos — práctica en el hogar</div>', unsafe_allow_html=True)
        lavado_manos = st.selectbox("", ["Siempre (antes de comer y después de ir al baño)","Frecuentemente","A veces","Casi nunca o nunca"], key="lav", label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    # ── 5. ANTECEDENTES ───────────────────────────────────────────────────────
    st.markdown("""<div class="section-card">
    <div class="section-card-title">5. Antecedentes personales y familiares</div>
    <div class="section-card-subtitle">Condiciones de salud relevantes que influyen en el estado nutricional</div>""", unsafe_allow_html=True)
    c1,c2 = st.columns(2)
    with c1:
        st.markdown('<div class="campo-label">Enfermedades crónicas personales</div>', unsafe_allow_html=True)
        enf_per = st.multiselect("", ["Diabetes mellitus tipo 1","Diabetes mellitus tipo 2","Hipertensión arterial","Hipotiroidismo","Hipertiroidismo","Enfermedad renal crónica","Enfermedad cardiovascular","Cáncer","VIH/SIDA","Tuberculosis","Anemia crónica","EPOC","Síndrome de malabsorción","Enfermedad celíaca","Otra"], key="enp", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Hospitalización previa por desnutrición</div>', unsafe_allow_html=True)
        hosp = st.selectbox("", ["No","Sí — hace menos de 1 año","Sí — hace 1-3 años","Sí — hace más de 3 años"], key="hos", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Medicamentos actuales (separar con coma)</div>', unsafe_allow_html=True)
        medicamentos = st.text_area("", height=60, key="med", label_visibility="collapsed", placeholder="Ej: Metformina, Enalapril, Sulfato ferroso...")
    with c2:
        st.markdown('<div class="campo-label">Antecedentes familiares relevantes</div>', unsafe_allow_html=True)
        ant_fam = st.multiselect("", ["Diabetes mellitus","Hipertensión arterial","Obesidad","Enfermedades cardiovasculares","Cáncer","Desnutrición infantil","Enfermedad celíaca","Osteoporosis","Dislipidemia","Otra"], key="anf", label_visibility="collapsed")
        st.markdown('<div class="campo-label">Alergias e intolerancias alimentarias</div>', unsafe_allow_html=True)
        alergias = st.multiselect("", ["Gluten (celiaquía)","Lactosa","Proteína de leche de vaca","Mariscos y crustáceos","Huevo","Maní y frutos secos","Soya","Colorantes y aditivos artificiales","Ninguna conocida"], key="ale", label_visibility="collapsed")
    obs_ant = st.text_area("Observaciones adicionales sobre antecedentes", height=60, key="oant", placeholder="Información relevante adicional...")
    st.markdown("</div>", unsafe_allow_html=True)

    # ── 6. FRECUENCIA DE CONSUMO ──────────────────────────────────────────────
    st.markdown("""<div class="section-card">
    <div class="section-card-title">6. Frecuencia de consumo de alimentos</div>
    <div class="section-card-subtitle">Clasificado por grupos de alimentos según Guías Alimentarias Colombia — ICBF</div>""", unsafe_allow_html=True)
    st.caption("Indique con qué frecuencia consume alimentos de cada grupo")

    freq_consumo = {}
    for grupo, alimentos in GRUPOS_ALIMENTOS.items():
        with st.expander(f"Grupo: {grupo}", expanded=False):
            st.caption(f"Alimentos incluidos en este grupo: {', '.join(alimentos)}")
            freq_consumo[grupo] = st.selectbox(
                f"Frecuencia de consumo — {grupo}",
                FRECUENCIAS_CONSUMO,
                key=f"fr_{grupo[:15]}"
            )

    obs_alim = st.text_area("Observaciones sobre hábitos alimentarios", height=60, key="oalim", placeholder="Preferencias, aversiones, restricciones culturales o religiosas...")
    st.markdown("</div>", unsafe_allow_html=True)

    # ── CALCULAR ──────────────────────────────────────────────────────────────
    if st.button("Generar evaluación nutricional completa"):
        imc = calcular_imc(peso, talla)
        edad_total_m = (edad_a * 12 + (edad_m_ad if es_menor else 0))

        if es_menor:
            clf, nivel, msg = clasificar_nino(imc, edad_total_m)
        else:
            clf, nivel, msg = clasificar_adulto(imc)

        recs = recomendaciones(clf)

        st.markdown(f'<div class="grupo-etario-badge">{grupo_sel}</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Resultados antropométricos</div>', unsafe_allow_html=True)

        if es_menor:
            m1,m2,m3,m4 = st.columns(4)
            m1.metric("IMC", f"{imc}")
            m2.metric("Edad", f"{edad_a}a {edad_m_ad}m", f"{edad_total_m} meses")
            m3.metric("Peso", f"{peso} kg")
            m4.metric("Talla", f"{talla} cm")
        else:
            pi = peso_ideal(talla, sexo)
            ge = get_energia(peso, talla, edad_a, sexo, actividad)
            m1,m2,m3,m4 = st.columns(4)
            m1.metric("IMC", f"{imc}", f"{'Normal' if 18.5<=imc<25 else 'Fuera de rango'}")
            m2.metric("Peso ideal", f"{pi} kg", f"{peso-pi:+.1f} kg")
            m3.metric("GET estimado", f"{ge} kcal", actividad)
            m4.metric("Talla", f"{talla} cm")

        if p_cintura > 0 and p_cadera > 0:
            icc = round(p_cintura/p_cadera, 2)
            ricc = "Alto" if (sexo=="Masculino" and icc>0.95) or (sexo=="Femenino" and icc>0.85) else "Normal"
            st.metric("Índice cintura/cadera", f"{icc}", f"Riesgo cardiovascular: {ricc}")

        css = {"normal":"result-normal","riesgo":"result-riesgo","malo":"result-malo","sobrepeso":"result-sobrepeso"}.get(nivel,"result-riesgo")
        st.markdown(f'<div class="result-card {css}"><h3>{clf}</h3><p>{msg}</p></div>', unsafe_allow_html=True)

        # Alertas automáticas
        if nivel == "malo":
            st.markdown('<div class="alerta-critica">⚠ ALERTA CLÍNICA — Intervención prioritaria. Notificar a SISVAN y coordinar con equipo de salud.</div>', unsafe_allow_html=True)
        if inseguridad in ["Inseguridad moderada","Inseguridad severa"]:
            st.markdown(f'<div class="alerta-critica">⚠ {inseguridad.upper()} — Vincular urgentemente a programas de asistencia alimentaria ICBF.</div>', unsafe_allow_html=True)
        if agua in ["Agua de río, lluvia o sin tratamiento","Sin acceso a agua segura"]:
            st.markdown('<div class="alerta-moderada">⚠ AGUA NO SEGURA — Riesgo de enfermedades diarreicas y afectación del aprovechamiento biológico de nutrientes.</div>', unsafe_allow_html=True)
        if despar in ["Hace más de 1 año","No recuerda","Nunca"]:
            st.markdown('<div class="alerta-moderada">⚠ DESPARASITACIÓN PENDIENTE — Programar según protocolo MinSalud. La parasitosis afecta el aprovechamiento biológico de nutrientes.</div>', unsafe_allow_html=True)
        if hosp != "No":
            st.markdown('<div class="alerta-moderada">⚠ ANTECEDENTE DE HOSPITALIZACIÓN POR DESNUTRICIÓN — Seguimiento prioritario y refuerzo del plan de recuperación.</div>', unsafe_allow_html=True)

        # Perfil de consumo
        st.markdown('<div class="section-title">Perfil de consumo alimentario por grupos</div>', unsafe_allow_html=True)
        problemas_consumo = []
        c1,c2 = st.columns(2)
        for i,(grupo,freq) in enumerate(freq_consumo.items()):
            if freq == "Diario": badge="freq-diario"
            elif "semana" in freq: badge="freq-semanal"
            elif freq in ["Quincenal","Mensual"]: badge="freq-quincenal"
            elif freq == "Ocasional": badge="freq-ocasional"
            else: badge="freq-nunca"
            with c1 if i%2==0 else c2:
                st.markdown(f'<div class="recom-item" style="padding:10px 14px;"><strong>{grupo}</strong><br><span class="freq-badge {badge}">{freq}</span></div>', unsafe_allow_html=True)
            if freq in ["Nunca","Ocasional","Mensual"] and "Azúcares" not in grupo and "Comidas rápidas" not in grupo:
                problemas_consumo.append(grupo)
        if problemas_consumo:
            st.markdown(f'<div class="alerta-moderada">⚠ BAJO CONSUMO detectado en: {", ".join(problemas_consumo)}. Incluir estrategias de mejora en el plan nutricional.</div>', unsafe_allow_html=True)

        # Plan de intervención
        st.markdown('<div class="section-title">Plan de intervención recomendado</div>', unsafe_allow_html=True)
        for t,d in recs:
            st.markdown(f'<div class="recom-item"><strong>{t}</strong> — {d}</div>', unsafe_allow_html=True)
        if alergias and "Ninguna conocida" not in alergias:
            st.markdown(f'<div class="recom-item"><strong>Alergias / intolerancias</strong> — Excluir del plan alimentario: {", ".join(alergias)}</div>', unsafe_allow_html=True)

        # Exportar Excel
        st.markdown('<div class="section-title">Exportar evaluación completa</div>', unsafe_allow_html=True)
        exp = {
            "Fecha":[str(fecha_ev)],"Grupo etario":[grupo_sel],"Norma":[info_grupo["norma"]],
            "Nombre":[nombre],"Documento":[doc],"Departamento":[depto],"Municipio":[mun],
            "Sexo":[sexo],"Edad (años)":[edad_a],"Peso (kg)":[peso],"Talla (cm)":[talla],"IMC":[imc],
            "Clasificación nutricional":[clf],
            "Estrato":[estrato],"Régimen salud":[regimen],"Tipo vivienda":[vivienda],
            "Personas hogar":[personas],"Escolaridad":[escolaridad],"Etnia":[etnia],
            "Comidas/día":[comidas],
            "Disponibilidad alimentos":[disponibilidad],"Fuente alimentos":[fuente_alimentos],
            "Acceso económico":[acceso_eco],"Acceso físico":[acceso_fisico],
            "Programas alimentarios":[", ".join(prog_alim) if prog_alim else "Ninguno"],
            "Decisión alimentaria hogar":[decision_alim],
            "Inseguridad alimentaria ELCSA":[inseguridad],
            "Agua potable":[agua],"Almacenamiento alimentos":[almacenamiento],
            "Preparación alimentos":[preparacion],
            "Condición que afecta absorción":[salud_absorcion],
            "Desparasitación":[despar],"Saneamiento básico":[saneamiento],
            "Lavado de manos":[lavado_manos],
            "Enfermedades personales":[", ".join(enf_per) if enf_per else "Ninguna"],
            "Antecedentes familiares":[", ".join(ant_fam) if ant_fam else "Ninguno"],
            "Hospitalización previa":[hosp],"Medicamentos":[medicamentos],
            "Alergias":[", ".join(alergias) if alergias else "Ninguna"],
            "Obs. antecedentes":[obs_ant],"Obs. alimentarias":[obs_alim],
        }
        for grupo,freq in freq_consumo.items():
            exp[f"Consumo — {grupo}"] = [freq]
        buf = io.BytesIO()
        pd.DataFrame(exp).to_excel(buf, index=False)
        st.download_button(
            "Descargar evaluación completa en Excel",
            data=buf.getvalue(),
            file_name=f"NutriVida_{nombre.replace(' ','_') if nombre else 'paciente'}_{grupo_sel[:15]}_{fecha_ev}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 2 — REFERENCIA EPIDEMIOLÓGICA (igual que v4, se mantiene)
# ═══════════════════════════════════════════════════════════════════════════════
elif modulo == "Referencia Epidemiológica":
    st.markdown("""<div style="background:linear-gradient(135deg,#2d1b69,#4a2c9e);border-radius:12px;padding:20px 28px;margin-bottom:20px;">
    <div style="font-family:'Playfair Display',serif;font-size:1.3rem;font-weight:700;color:white;margin-bottom:4px;">Panel de Referencia Epidemiológica Nacional</div>
    <div style="font-size:0.78rem;color:rgba(255,255,255,0.6);font-family:'DM Mono',monospace;">ENSIN 2015 · SIVIGILA · MinSalud Colombia · Indicadores de salud pública</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div style="background:linear-gradient(135deg,#1a3a2a,#2d6a4f);border-radius:10px;
                padding:10px 20px;margin-bottom:16px;">
        <div style="font-size:0.7rem;color:rgba(255,255,255,0.6);font-family:'DM Mono',monospace;
                    letter-spacing:1px;">Selecciona una sección del panel epidemiológico</div>
    </div>
    """, unsafe_allow_html=True)
    etabs = st.tabs([
        "📊  ENSIN 2015 — Indicadores",
        "🏥  Salud pública y cronicidad",
        "📈  Tasas y prevalencias SIVIGILA",
        "⚖️  Doble carga nutricional",
        "🗺️  Comparador departamental"
    ])

    with etabs[0]:
        st.markdown('<div class="section-title">Indicadores ENSIN 2015 — Colombia</div>', unsafe_allow_html=True)
        st.markdown('<div class="recom-item"><strong>Fuente:</strong> ENSIN 2015 — ICBF, MinSalud, INS, Profamilia. <a href="https://www.icbf.gov.co/bienestar/nutricion/encuesta-nacional-situacion-nutricional" target="_blank">Ver fuente oficial →</a></div>', unsafe_allow_html=True)
        st.dataframe(pd.DataFrame({"Indicador":["LM exclusiva < 6 meses","LM continua 6-11 meses","LM continua al 2° año","Inicio temprano LM","Desnutrición crónica < 5 años","Desnutrición global < 5 años","Desnutrición aguda < 5 años","Sobrepeso adultos","Obesidad adultos","Anemia niños 6-23 m"],"Colombia ENSIN 2015":["36.1%","66.3%","32.4%","72.5%","10.8%","3.7%","2.3%","37.7%","18.7%","27.7%"],"Meta OMS/MinSalud":[">50%",">80%",">60%",">80%","<5%","<3%","<2%","<30%","<10%","<10%"],"Tendencia":["↓","↑","↓","↑","↓","↓","↓","↑","↑","↓"]}), use_container_width=True, hide_index=True)

        c1,c2 = st.columns(2)
        with c1:
            st.caption("Estado nutricional niños < 5 años")
            st.dataframe(pd.DataFrame({"Indicador":["Desnutrición crónica","Desnutrición global","Desnutrición aguda","Riesgo sobrepeso","Sobrepeso y obesidad"],"Prevalencia":["10.8%","3.7%","2.3%","4.9%","6.3%"],"Vs 2010":["↓","↓","↓","↑","↑"]}), use_container_width=True, hide_index=True)
        with c2:
            st.caption("Estado nutricional adultos 18-64 años")
            st.dataframe(pd.DataFrame({"Indicador":["Delgadez (IMC<18.5)","Normal","Sobrepeso","Obesidad","Obesidad abdominal"],"Prevalencia":["3.1%","47.5%","37.7%","18.7%","53.6%"],"Vs 2010":["↓","↓","↑","↑","↑"]}), use_container_width=True, hide_index=True)

        st.markdown('<div class="section-title">Deficiencia de micronutrientes</div>', unsafe_allow_html=True)
        c1,c2,c3 = st.columns(3)
        with c1:
            st.caption("Anemia")
            st.dataframe(pd.DataFrame({"Grupo":["Niños 6-23 m","Niños 2-4 a","Mujeres 13-49 a","Gestantes","Adultos ≥65 a"],"Prev.":["27.7%","14.8%","10.8%","17.9%","12.4%"]}), use_container_width=True, hide_index=True)
        with c2:
            st.caption("Vitamina A")
            st.dataframe(pd.DataFrame({"Grupo":["Niños 1-4 a","Niños 5-12 a","Gestantes","Lactantes"],"Def.":["24.3%","8.5%","3.1%","5.2%"]}), use_container_width=True, hide_index=True)
        with c3:
            st.caption("Zinc")
            st.dataframe(pd.DataFrame({"Grupo":["Niños 1-4 a","Niños 5-12 a","Adolescentes","Adultos"],"Def.":["43.3%","28.7%","18.4%","12.1%"]}), use_container_width=True, hide_index=True)

    with etabs[1]:
        st.markdown('<div class="section-title">Envejecimiento poblacional — Colombia</div>', unsafe_allow_html=True)
        st.markdown('<div class="recom-item"><strong>Fuente:</strong> <a href="https://www.dane.gov.co/index.php/estadisticas-por-tema/demografia-y-poblacion/proyecciones-de-poblacion" target="_blank">DANE Proyecciones 2018-2050</a> · <a href="https://www.minsalud.gov.co/proteccionsocial/Paginas/envejecimiento-vejez.aspx" target="_blank">MinSalud Política de Envejecimiento</a></div>', unsafe_allow_html=True)
        c1,c2 = st.columns(2)
        with c1:
            st.dataframe(pd.DataFrame({"Indicador":["Índice de envejecimiento","Razón dependencia total","Razón dependencia AM","Población ≥60 años (2023)","Población ≥60 años (2050)","Esperanza de vida al nacer","Años de vida saludable (AVISA)"],"Colombia":["58.3","50.2","11.8","11.4%","23.8%","77.3 años","67.1 años"],"Ref. OPS/LAC":["75.0","52.1","14.2","11.8%","25.1%","75.8 años","65.8 años"]}), use_container_width=True, hide_index=True)
        with c2:
            fig_p=go.Figure()
            fig_p.add_trace(go.Bar(name="2023",x=[7.8,15.2,22.1,20.3,18.4,11.4,4.8],y=["0-4","5-14","15-29","30-44","45-59","60-74","75+"],orientation="h",marker_color="#a855f7"))
            fig_p.add_trace(go.Bar(name="2050 proy.",x=[5.9,11.8,17.2,19.1,19.8,17.6,8.6],y=["0-4","5-14","15-29","30-44","45-59","60-74","75+"],orientation="h",marker_color="#f59e0b",opacity=0.7))
            fig_p.update_layout(barmode="group",height=250,margin=dict(l=0,r=0,t=20,b=0),title=dict(text="Distribución etaria (%)",font=dict(size=11,color="#9580c0")),paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",xaxis=dict(color="#9580c0",gridcolor="#f0e8ff",title="%"),yaxis=dict(color="#4a3570"),legend=dict(font=dict(color="#4a3570",size=10),bgcolor="rgba(0,0,0,0)"))
            st.plotly_chart(fig_p, use_container_width=True)

        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Esp. vida mujeres","80.1 años","↑ vs 2015")
        m2.metric("Esp. vida hombres","74.2 años","↑ vs 2015")
        m3.metric("AVISA","67.1 años","OPS: 65.8 a")
        m4.metric("Años perdidos ENT","13.2 años","Principal: dieta")

        st.markdown('<div class="section-title">ENT y cronicidad relacionada con nutrición</div>', unsafe_allow_html=True)
        st.markdown('<div class="recom-item"><strong>Fuente:</strong> <a href="https://www.minsalud.gov.co/salud/publica/PENT/Paginas/enfermedades-no-transmisibles.aspx" target="_blank">MinSalud — Enfermedades No Transmisibles</a></div>', unsafe_allow_html=True)
        c1,c2 = st.columns(2)
        with c1:
            st.dataframe(pd.DataFrame({"ENT":["Hipertensión arterial","Diabetes mellitus tipo 2","Obesidad","Enfermedad cardiovascular","EPOC","Síndrome metabólico","Dislipidemia"],"Prevalencia":["22.8%","8.1%","18.7%","6.3%","8.9%","31.4%","45.2%"],"Rel. nutrición":["Alta","Alta","Directa","Alta","Moderada","Directa","Alta"]}), use_container_width=True, hide_index=True)
        with c2:
            fig_e=go.Figure(go.Bar(x=[22.8,8.1,18.7,6.3,8.9,31.4],y=["HTA","Diabetes","Obesidad","ECV","EPOC","S. metabólico"],orientation="h",marker=dict(color=[22.8,8.1,18.7,6.3,8.9,31.4],colorscale=[[0,"#e9d5ff"],[0.5,"#a855f7"],[1,"#2d1b69"]])))
            fig_e.update_layout(height=250,margin=dict(l=0,r=0,t=10,b=0),paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",xaxis=dict(color="#9580c0",gridcolor="#f0e8ff",title="%"),yaxis=dict(color="#4a3570"))
            st.plotly_chart(fig_e, use_container_width=True)

    with etabs[2]:
        st.markdown('<div class="section-title">Tasas y prevalencias — SIVIGILA</div>', unsafe_allow_html=True)
        st.markdown('<div class="recom-item"><strong>Fuente:</strong> <a href="https://www.ins.gov.co/Noticias/Paginas/sivigila.aspx" target="_blank">SIVIGILA — INS Colombia</a> · <a href="https://www.minsalud.gov.co/salud/publica/epidemiologia/Paginas/boletin-epidemiologico.aspx" target="_blank">Boletines epidemiológicos MinSalud</a></div>', unsafe_allow_html=True)
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("Desnutrición aguda < 5 a","2.3%","ENSIN 2015")
        m2.metric("Mortalidad desnutrición","3.2 x 100k","↓ vs 2015")
        m3.metric("Anemia niños 6-23 m","27.7%","Más vulnerable")
        m4.metric("Cobertura supl. Fe gestantes","68.4%","Meta: 90%")

        df_dd=pd.DataFrame({"Departamento":["La Guajira","Chocó","Vaupés","Vichada","Guainía","Amazonas","Córdoba","Magdalena","Sucre","Bolívar","Nariño","Cauca","Colombia"],"Prev. desnut. aguda (%)": [10.1,7.8,6.9,6.4,5.8,5.2,4.1,3.8,3.4,3.1,2.8,2.6,2.3],"Casos SIVIGILA 2023":[2841,1923,412,318,287,241,1876,1654,1432,1387,1243,1187,28940],"Riesgo":["Crítico","Crítico","Alto","Alto","Alto","Alto","Moderado","Moderado","Moderado","Moderado","Moderado","Moderado","Referencia"]})
        st.dataframe(df_dd, use_container_width=True, hide_index=True)
        fig_dd=px.bar(df_dd.head(12),x="Prev. desnut. aguda (%)",y="Departamento",orientation="h",color="Prev. desnut. aguda (%)",color_continuous_scale=[[0,"#fef9c3"],[0.4,"#f59e0b"],[0.7,"#ef4444"],[1,"#7f1d1d"]])
        fig_dd.update_layout(height=300,margin=dict(l=0,r=0,t=10,b=0),paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",xaxis=dict(color="#9580c0",gridcolor="#f0e8ff"),yaxis=dict(color="#4a3570",autorange="reversed"),coloraxis_showscale=False)
        st.plotly_chart(fig_dd, use_container_width=True)

        c1,c2=st.columns(2)
        with c1:
            df_an=pd.DataFrame({"Grupo":["Niños 6-23 m","Niños 2-4 a","Niños 5-12 a","Adolescentes","Mujeres 18-49 a","Gestantes","Adultos ≥65 a"],"Prevalencia (%)": [27.7,14.8,7.3,6.1,10.8,17.9,12.4]})
            fig_an=px.bar(df_an,x="Grupo",y="Prevalencia (%)",color="Prevalencia (%)",color_continuous_scale=[[0,"#fef2f2"],[0.5,"#f87171"],[1,"#b91c1c"]])
            fig_an.update_layout(height=240,margin=dict(l=0,r=0,t=20,b=0),title=dict(text="Anemia por grupo de edad",font=dict(size=11,color="#9580c0")),paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",xaxis=dict(color="#9580c0",tickangle=-30,tickfont=dict(size=9)),yaxis=dict(color="#9580c0",gridcolor="#f0e8ff"),coloraxis_showscale=False)
            st.plotly_chart(fig_an, use_container_width=True)
        with c2:
            df_ob=pd.DataFrame({"Departamento":["San Andrés","Atlántico","Valle del Cauca","Risaralda","Bogotá","Antioquia","Santander","Cundinamarca","Nariño","Chocó"],"Obesidad adultos (%)": [26.4,24.1,22.8,21.9,20.4,19.8,18.9,18.1,13.2,11.8]})
            fig_ob=px.bar(df_ob,x="Obesidad adultos (%)",y="Departamento",orientation="h",color="Obesidad adultos (%)",color_continuous_scale=[[0,"#faf5ff"],[0.5,"#a855f7"],[1,"#2d1b69"]])
            fig_ob.update_layout(height=240,margin=dict(l=0,r=0,t=20,b=0),title=dict(text="Obesidad adultos por departamento",font=dict(size=11,color="#9580c0")),paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",xaxis=dict(color="#9580c0",gridcolor="#f0e8ff"),yaxis=dict(color="#4a3570",autorange="reversed"),coloraxis_showscale=False)
            st.plotly_chart(fig_ob, use_container_width=True)

        df_sup=pd.DataFrame({"Programa":["Hierro-ácido fólico gestantes","Vitamina A niños 6-59 m","Hierro niños 6-23 m","Calcio gestantes","Desparasitación 12-59 m","Complementación ICBF"],"Cobertura (%)": [68.4,71.2,52.3,44.8,61.7,78.9],"Meta MinSalud (%)": [90,85,80,70,80,90]})
        fig_sup=go.Figure()
        fig_sup.add_trace(go.Bar(name="Cobertura actual",x=df_sup["Cobertura (%)"],y=df_sup["Programa"],orientation="h",marker_color="#a855f7"))
        fig_sup.add_trace(go.Bar(name="Meta MinSalud",x=df_sup["Meta MinSalud (%)"],y=df_sup["Programa"],orientation="h",marker_color="#e9d5ff",opacity=0.5))
        fig_sup.update_layout(barmode="overlay",height=260,margin=dict(l=0,r=0,t=20,b=0),title=dict(text="Cobertura suplementación vs meta",font=dict(size=11,color="#9580c0")),paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",xaxis=dict(color="#9580c0",gridcolor="#f0e8ff",title="%"),yaxis=dict(color="#4a3570",tickfont=dict(size=9)),legend=dict(font=dict(color="#4a3570",size=10),bgcolor="rgba(0,0,0,0)"))
        st.plotly_chart(fig_sup, use_container_width=True)

    with etabs[3]:
        st.markdown('<div class="section-title">Doble carga nutricional</div>', unsafe_allow_html=True)
        st.markdown('<div class="recom-item"><strong>¿Qué es?</strong> — Coexistencia de desnutrición y exceso de peso en una misma población, hogar o individuo. Colombia enfrenta esta paradoja nutricional especialmente en zonas de alta pobreza. <a href="https://www.fao.org/colombia/noticias/detail-events/es/c/1157928/" target="_blank">Ver más →</a></div>', unsafe_allow_html=True)
        m1,m2=st.columns(2)
        m1.metric("Hogares con doble carga","11.4%","ENSIN 2015")
        m2.metric("Niños desnutridos con madre con sobrepeso","28.3%","Paradoja nutricional Colombia")
        c1,c2=st.columns(2)
        with c1:
            df_dc=pd.DataFrame({"Departamento":["La Guajira","Chocó","Córdoba","Bolívar","Nariño","Cauca","Valle del Cauca","Antioquia","Cundinamarca","Atlántico"],"Desnut. crónica (%)": [27.9,18.2,12.4,11.8,16.3,15.1,6.2,7.8,8.1,6.9],"Obesidad adultos (%)": [9.6,11.8,14.2,13.7,13.2,12.8,22.8,19.8,18.1,24.1]})
            fig_dc=go.Figure(go.Scatter(x=df_dc["Obesidad adultos (%)"],y=df_dc["Desnut. crónica (%)"],mode="markers+text",text=df_dc["Departamento"],textposition="top center",textfont=dict(size=8,color="#4a3570"),marker=dict(size=12,color="#a855f7",opacity=0.7)))
            fig_dc.update_layout(height=280,margin=dict(l=0,r=0,t=20,b=0),paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",xaxis=dict(title="Obesidad adultos (%)",color="#9580c0",gridcolor="#f0e8ff"),yaxis=dict(title="Desnut. crónica niños (%)",color="#9580c0",gridcolor="#f0e8ff"),title=dict(text="Doble carga por departamento",font=dict(size=11,color="#9580c0")))
            st.plotly_chart(fig_dc, use_container_width=True)
        with c2:
            for t,d in [("Transición nutricional","Cambio de dietas tradicionales a ultraprocesados de bajo costo"),("Inseguridad alimentaria","Hogares que priorizan cantidad sobre calidad nutricional"),("Urbanización","Pérdida de producción propia de alimentos frescos"),("Publicidad alimentos","Exposición a ultraprocesados en población vulnerable"),("Pobreza","Limita acceso a alimentos de calidad nutricional")]:
                st.markdown(f'<div class="recom-item" style="padding:9px 14px;"><strong>{t}</strong> — {d}</div>', unsafe_allow_html=True)
        años=["ENSIN 2005","ENSIN 2010","ENSIN 2015","Proy. 2025"]
        fig_t=go.Figure()
        fig_t.add_trace(go.Scatter(x=años,y=[13.2,13.2,10.8,8.5],name="Desnut. crónica < 5 años (%)",line=dict(color="#ef4444",width=2.5),mode="lines+markers"))
        fig_t.add_trace(go.Scatter(x=años,y=[45.9,51.2,56.4,62.1],name="Exceso de peso adultos (%)",line=dict(color="#a855f7",width=2.5),mode="lines+markers"))
        fig_t.update_layout(height=220,margin=dict(l=0,r=0,t=10,b=0),paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",xaxis=dict(color="#9580c0",gridcolor="#f0e8ff"),yaxis=dict(color="#9580c0",gridcolor="#f0e8ff",title="%"),legend=dict(font=dict(color="#4a3570",size=10),bgcolor="rgba(0,0,0,0)"))
        st.plotly_chart(fig_t, use_container_width=True)

    with etabs[4]:
        st.markdown('<div class="section-title">Comparador departamental</div>', unsafe_allow_html=True)
        IND_D={"La Guajira":{"dc":27.9,"da":10.1,"ob":9.6,"an":38.2,"ia":58.3,"cs":41.2},"Chocó":{"dc":18.2,"da":7.8,"ob":11.8,"an":32.1,"ia":52.1,"cs":48.7},"Vaupés":{"dc":15.4,"da":6.9,"ob":8.4,"an":29.8,"ia":48.6,"cs":44.1},"Vichada":{"dc":14.8,"da":6.4,"ob":10.4,"an":28.4,"ia":46.2,"cs":45.8},"Nariño":{"dc":16.3,"da":2.8,"ob":13.2,"an":24.1,"ia":38.4,"cs":62.3},"Cauca":{"dc":15.1,"da":2.6,"ob":12.8,"an":22.8,"ia":36.7,"cs":64.1},"Córdoba":{"dc":12.4,"da":4.1,"ob":14.2,"an":21.4,"ia":34.2,"cs":61.8},"Bolívar":{"dc":11.8,"da":3.1,"ob":13.7,"an":20.1,"ia":32.8,"cs":63.2},"Magdalena":{"dc":10.2,"da":3.8,"ob":12.4,"an":18.9,"ia":30.1,"cs":65.4},"Antioquia":{"dc":7.8,"da":1.8,"ob":19.8,"an":14.2,"ia":21.4,"cs":72.1},"Cundinamarca":{"dc":8.1,"da":1.4,"ob":18.1,"an":13.8,"ia":18.9,"cs":74.8},"Valle del Cauca":{"dc":6.2,"da":1.2,"ob":22.8,"an":11.4,"ia":16.2,"cs":76.4},"Atlántico":{"dc":6.9,"da":1.1,"ob":24.1,"an":12.1,"ia":17.8,"cs":75.2},"Bogotá D.C.":{"dc":6.1,"da":0.9,"ob":20.4,"an":10.8,"ia":14.1,"cs":78.9},"San Andrés":{"dc":4.2,"da":0.7,"ob":26.4,"an":9.4,"ia":12.3,"cs":80.1},"Nacional":{"dc":10.8,"da":2.3,"ob":18.7,"an":27.7,"ia":54.2,"cs":68.4}}
        ds=st.selectbox("Seleccionar departamento",list(IND_D.keys()))
        ind=IND_D[ds]; nac=IND_D["Nacional"]
        k1,k2,k3=st.columns(3)
        k1.metric("Desnutrición crónica",f"{ind['dc']}%",f"{ind['dc']-nac['dc']:+.1f}pp vs nac.")
        k2.metric("Desnutrición aguda",f"{ind['da']}%",f"{ind['da']-nac['da']:+.1f}pp vs nac.")
        k3.metric("Obesidad adultos",f"{ind['ob']}%",f"{ind['ob']-nac['ob']:+.1f}pp vs nac.")
        k4,k5,k6=st.columns(3)
        k4.metric("Anemia niños 6-23 m",f"{ind['an']}%",f"{ind['an']-nac['an']:+.1f}pp vs nac.")
        k5.metric("Inseguridad alimentaria",f"{ind['ia']}%",f"{ind['ia']-nac['ia']:+.1f}pp vs nac.")
        k6.metric("Cobertura suplementación",f"{ind['cs']}%",f"{ind['cs']-nac['cs']:+.1f}pp vs nac.")
        cats=["Desnut.\ncrónica","Desnut.\naguda","Obesidad","Anemia","Inseg.\naliment.","Cobert.\nsupl."]
        vd=[ind["dc"],ind["da"],ind["ob"],ind["an"],ind["ia"],ind["cs"]]
        vn=[nac["dc"],nac["da"],nac["ob"],nac["an"],nac["ia"],nac["cs"]]
        fig_r=go.Figure()
        fig_r.add_trace(go.Scatterpolar(r=vd+[vd[0]],theta=cats+[cats[0]],fill="toself",name=ds,line_color="#a855f7",fillcolor="rgba(168,85,247,0.15)"))
        fig_r.add_trace(go.Scatterpolar(r=vn+[vn[0]],theta=cats+[cats[0]],fill="toself",name="Nacional",line_color="#f59e0b",fillcolor="rgba(245,158,11,0.1)"))
        fig_r.update_layout(polar=dict(radialaxis=dict(visible=True,color="#9580c0"),angularaxis=dict(color="#4a3570")),height=320,margin=dict(l=20,r=20,t=30,b=20),paper_bgcolor="rgba(0,0,0,0)",legend=dict(font=dict(color="#4a3570",size=10),bgcolor="rgba(0,0,0,0)"))
        st.plotly_chart(fig_r, use_container_width=True)
        score=sum([3 if ind["dc"]>15 else 2 if ind["dc"]>10 else 1,3 if ind["ia"]>45 else 2 if ind["ia"]>25 else 1,2 if ind["cs"]<55 else 1 if ind["cs"]<70 else 0])
        if score>=7: cls,css2="RIESGO CRÍTICO","alerta-critica"
        elif score>=5: cls,css2="RIESGO ALTO","alerta-moderada"
        else: cls,css2="RIESGO MODERADO/BAJO","recom-item"
        st.markdown(f'<div class="{css2}"><strong>{ds} — {cls}</strong> — Puntuación de vulnerabilidad: {score}/8</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 3 — DASHBOARD POBLACIONAL AVANZADO v6.0
# ═══════════════════════════════════════════════════════════════════════════════
elif modulo == "Dashboard Poblacional":

    # Coordenadas centroides departamentos Colombia
    COORD_DEPTOS = {
        "Amazonas":(-1.44,-71.57),"Antioquia":(7.19,-75.34),"Arauca":(6.54,-71.00),
        "Atlántico":(10.69,-74.87),"Bolívar":(8.67,-74.03),"Boyacá":(5.45,-73.36),
        "Caldas":(5.30,-75.27),"Caquetá":(0.87,-73.84),"Casanare":(5.76,-71.57),
        "Cauca":(2.44,-76.61),"Cesar":(9.33,-73.50),"Chocó":(5.69,-76.65),
        "Córdoba":(8.38,-75.88),"Cundinamarca":(4.60,-74.08),"Guainía":(2.57,-68.52),
        "Guaviare":(2.04,-72.33),"Huila":(2.53,-75.52),"La Guajira":(11.35,-72.48),
        "Magdalena":(10.41,-74.40),"Meta":(3.99,-73.56),"Nariño":(1.28,-77.35),
        "Norte de Santander":(7.94,-72.49),"Putumayo":(0.43,-76.64),
        "Quindío":(4.46,-75.67),"Risaralda":(5.31,-75.99),
        "San Andrés y Providencia":(12.53,-81.72),"Santander":(6.64,-73.65),
        "Sucre":(9.30,-75.40),"Tolima":(3.81,-75.19),"Valle del Cauca":(3.80,-76.52),
        "Vaupés":(0.85,-70.81),"Vichada":(4.42,-69.67),
    }

    # ── Datos demo realistas ──────────────────────────────────────────────────
    @st.cache_data
    def generar_datos_avanzados():
        np.random.seed(42)
        n = 1250
        deptos = list(DEPARTAMENTOS.keys())
        pesos = np.array([0.03,0.12,0.02,0.05,0.05,0.03,0.02,0.02,0.02,
               0.03,0.03,0.04,0.04,0.08,0.01,0.01,0.03,0.05,
               0.03,0.03,0.03,0.03,0.02,0.02,0.02,0.01,0.04,
               0.02,0.03,0.05,0.01,0.01])
        pesos = pesos[:len(deptos)]
        pesos = pesos / pesos.sum()
        deptos_arr = np.random.choice(deptos, n, p=pesos)
        municipios = [np.random.choice(DEPARTAMENTOS[d]) for d in deptos_arr]
        sexos = np.random.choice(["Masculino","Femenino"], n)
        grupos_e = np.random.choice(list(GRUPOS_ETARIOS.keys()), n,
            p=[0.20,0.18,0.15,0.15,0.20,0.12])
        edades = []
        for g in grupos_e:
            a,b = GRUPOS_ETARIOS[g]["rango"]
            edades.append(np.random.randint(max(0,a*12), min(b*12+12, 1440)))
        edades = np.array(edades)
        # Estado nutricional realista por departamento
        riesgo_dpto = {
            "La Guajira":0.45,"Chocó":0.38,"Vaupés":0.35,"Vichada":0.32,"Guainía":0.30,
            "Amazonas":0.28,"Nariño":0.25,"Cauca":0.24,"Córdoba":0.22,"Bolívar":0.20,
            "Magdalena":0.19,"Sucre":0.18,"Huila":0.15,"Caquetá":0.14,"Arauca":0.13,
            "Casanare":0.12,"Boyacá":0.11,"Caldas":0.10,"Quindío":0.09,"Risaralda":0.09,
            "Tolima":0.10,"Meta":0.10,"Putumayo":0.13,"Norte de Santander":0.11,
            "Santander":0.09,"Cundinamarca":0.08,"Antioquia":0.09,"Atlántico":0.10,
            "Valle del Cauca":0.09,"Guaviare":0.15,"Cesar":0.16,"San Andrés y Providencia":0.07,
        }
        estados = []
        for d in deptos_arr:
            r = riesgo_dpto.get(d, 0.15)
            p = [r*0.3, r*0.4, r*0.3, 0.45-r*0.3, 0.15, 0.10]
            p = np.array(p); p = p/p.sum()
            estados.append(np.random.choice(
                ["Desnutrición severa","Desnutrición moderada","Desnutrición leve",
                 "Normal","Sobrepeso","Obesidad"], p=p))
        meses = pd.date_range("2024-01-01","2024-12-31",periods=n)
        inseg = np.where(np.isin(deptos_arr,["La Guajira","Chocó","Vaupés","Vichada"]),
            np.random.choice(["Inseguridad severa","Inseguridad moderada"],n),
            np.random.choice(["Seguridad alimentaria","Inseguridad leve","Inseguridad moderada"],n,p=[0.5,0.3,0.2]))
        agua = np.where(np.isin(deptos_arr,["La Guajira","Chocó","Vaupés"]),
            "Sin acceso a agua segura",
            np.random.choice(["Acueducto certificado","Agua comprada","Sin acceso a agua segura"],n,p=[0.6,0.3,0.1]))
        gestantes = np.random.choice([True,False],n,p=[0.08,0.92])
        control_nut = np.where(gestantes,
            np.random.choice([True,False],n,p=[0.55,0.45]), True)
        perdida_peso = np.random.choice([True,False],n,p=[0.10,0.90])
        return pd.DataFrame({
            "fecha":meses,"departamento":deptos_arr,"municipio":municipios,
            "sexo":sexos,"grupo_etario":grupos_e,"edad_meses":edades,
            "estado_nutricional":estados,"inseguridad_alimentaria":inseg,
            "agua_potable":agua,"gestante":gestantes,"control_nutricional":control_nut,
            "perdida_peso_reciente":perdida_peso,
        })

    # Cargar datos
    st.markdown("""
    <div style="background:linear-gradient(135deg,#2d1b69,#4a2c9e);border-radius:12px;
                padding:16px 24px;margin-bottom:20px;display:flex;align-items:center;gap:16px;">
        <div>
            <div style="font-family:'Playfair Display',serif;font-size:1.2rem;font-weight:700;color:white;">
                Dashboard Poblacional Avanzado
            </div>
            <div style="font-size:0.75rem;color:rgba(255,255,255,0.6);font-family:'DM Mono',monospace;">
                Vigilancia nutricional · Nacional · Tiempo real
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Opción: datos demo o datos reales
    st.markdown('<div class="section-title">Selecciona la fuente de datos</div>', unsafe_allow_html=True)
    c_r1, c_r2 = st.columns(2)
    with c_r1:
        st.markdown("""<div class="recom-item" style="border-left:3px solid #2d6a4f;padding:10px 14px;">
        <strong>Datos de demostración</strong> — 1.250 pacientes simulados con distribución
        realista por departamento. Ideal para explorar todas las funcionalidades del dashboard.
        </div>""", unsafe_allow_html=True)
    with c_r2:
        st.markdown("""<div class="recom-item" style="border-left:3px solid #f59e0b;padding:10px 14px;">
        <strong>Mis datos reales</strong> — Carga tu archivo Excel con datos reales de pacientes.
        El dashboard se actualiza automáticamente con tu información.
        </div>""", unsafe_allow_html=True)
    fuente_datos = st.radio("",
        ["Datos de demostración (1.250 pacientes simulados)","Cargar mis datos reales (Excel)"],
        horizontal=True, label_visibility="collapsed")

    if fuente_datos == "Cargar mis datos reales (Excel)":
        archivo_real = st.file_uploader("Cargar archivo Excel con datos reales", type=["xlsx","csv"])
        if archivo_real:
            try:
                df_raw = pd.read_csv(archivo_real) if archivo_real.name.endswith(".csv") else pd.read_excel(archivo_real)
                df = df_raw.copy()
                if "fecha" not in df.columns:
                    df["fecha"] = pd.Timestamp("2024-01-01")
                if "estado_nutricional" not in df.columns and "clasificacion" in df.columns:
                    df["estado_nutricional"] = df["clasificacion"]
                if "grupo_etario" not in df.columns:
                    df["grupo_etario"] = "No especificado"
                if "gestante" not in df.columns:
                    df["gestante"] = False
                if "perdida_peso_reciente" not in df.columns:
                    df["perdida_peso_reciente"] = False
                if "inseguridad_alimentaria" not in df.columns:
                    df["inseguridad_alimentaria"] = "No evaluada"
                if "agua_potable" not in df.columns:
                    df["agua_potable"] = "No evaluada"
                if "control_nutricional" not in df.columns:
                    df["control_nutricional"] = True
                st.success(f"✓ {len(df)} registros cargados correctamente")
            except Exception as e:
                st.error(f"Error: {e}")
                df = generar_datos_avanzados()
        else:
            st.info("Sube tu archivo o usa los datos de demostración")
            df = generar_datos_avanzados()
    else:
        df = generar_datos_avanzados()

    df["fecha"] = pd.to_datetime(df["fecha"])

    # ── FILTROS ───────────────────────────────────────────────────────────────
    with st.expander("Filtros de análisis", expanded=False):
        f1,f2,f3 = st.columns(3)
        with f1:
            st.markdown('<div class="campo-label">Departamento</div>', unsafe_allow_html=True)
            deptos_sel = st.multiselect("", sorted(df["departamento"].unique()),
                default=list(df["departamento"].unique()), key="df1", label_visibility="collapsed")
        with f2:
            st.markdown('<div class="campo-label">Grupo etario</div>', unsafe_allow_html=True)
            grupos_sel = st.multiselect("", list(GRUPOS_ETARIOS.keys()),
                default=list(GRUPOS_ETARIOS.keys()), key="df2", label_visibility="collapsed")
        with f3:
            st.markdown('<div class="campo-label">Período</div>', unsafe_allow_html=True)
            periodo = st.selectbox("", ["Todos","Último trimestre","Último semestre","Último año"],
                key="df3", label_visibility="collapsed")

    mask = df["departamento"].isin(deptos_sel if deptos_sel else df["departamento"].unique())
    if grupos_sel:
        mask = mask & df["grupo_etario"].isin(grupos_sel)
    if periodo == "Último trimestre":
        mask = mask & (df["fecha"] >= df["fecha"].max() - pd.Timedelta(days=90))
    elif periodo == "Último semestre":
        mask = mask & (df["fecha"] >= df["fecha"].max() - pd.Timedelta(days=180))
    dff = df[mask]
    total = len(dff)

    # ════════════════════════════════════════════════════════════════════════
    # PANEL 1 — KPIs PRINCIPALES
    # ════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="section-title">Panel principal — visión rápida</div>', unsafe_allow_html=True)

    desnut_n  = len(dff[dff["estado_nutricional"].isin(["Desnutrición severa","Desnutrición moderada","Desnutrición leve"])])
    riesgo_n  = len(dff[dff["estado_nutricional"].isin(["Desnutrición leve","Sobrepeso"])])
    normal_n  = len(dff[dff["estado_nutricional"]=="Normal"])
    exceso_n  = len(dff[dff["estado_nutricional"].isin(["Sobrepeso","Obesidad"])])
    severos_n = len(dff[dff["estado_nutricional"]=="Desnutrición severa"])

    pct = lambda n: f"{n/total*100:.1f}%" if total > 0 else "0%"

    k1,k2,k3,k4,k5 = st.columns(5)
    with k1:
        st.markdown(f"""<div style="background:white;border:1px solid #e8e0f5;border-radius:14px;
            padding:18px 16px;text-align:center;box-shadow:0 2px 8px rgba(45,27,105,0.06);">
            <div style="font-size:0.68rem;color:#9580c0;font-family:'DM Mono',monospace;
                        text-transform:uppercase;letter-spacing:1px;">Población evaluada</div>
            <div style="font-family:'Playfair Display',serif;font-size:2rem;color:#2d1b69;
                        font-weight:700;margin:6px 0;">{total:,}</div>
            <div style="font-size:0.72rem;color:#9580c0;">pacientes</div>
        </div>""", unsafe_allow_html=True)
    with k2:
        st.markdown(f"""<div style="background:#fef2f2;border:1px solid #fca5a5;border-radius:14px;
            padding:18px 16px;text-align:center;">
            <div style="font-size:0.68rem;color:#b91c1c;font-family:'DM Mono',monospace;
                        text-transform:uppercase;letter-spacing:1px;">🔴 Desnutrición</div>
            <div style="font-family:'Playfair Display',serif;font-size:2rem;color:#b91c1c;
                        font-weight:700;margin:6px 0;">{pct(desnut_n)}</div>
            <div style="font-size:0.72rem;color:#b91c1c;">{desnut_n:,} casos</div>
        </div>""", unsafe_allow_html=True)
    with k3:
        st.markdown(f"""<div style="background:#fffbeb;border:1px solid #fcd34d;border-radius:14px;
            padding:18px 16px;text-align:center;">
            <div style="font-size:0.68rem;color:#92400e;font-family:'DM Mono',monospace;
                        text-transform:uppercase;letter-spacing:1px;">🟠 Riesgo nutricional</div>
            <div style="font-family:'Playfair Display',serif;font-size:2rem;color:#b45309;
                        font-weight:700;margin:6px 0;">{pct(riesgo_n)}</div>
            <div style="font-size:0.72rem;color:#92400e;">{riesgo_n:,} casos</div>
        </div>""", unsafe_allow_html=True)
    with k4:
        st.markdown(f"""<div style="background:#f0faf4;border:1px solid #86efac;border-radius:14px;
            padding:18px 16px;text-align:center;">
            <div style="font-size:0.68rem;color:#15803d;font-family:'DM Mono',monospace;
                        text-transform:uppercase;letter-spacing:1px;">🟢 Estado normal</div>
            <div style="font-family:'Playfair Display',serif;font-size:2rem;color:#15803d;
                        font-weight:700;margin:6px 0;">{pct(normal_n)}</div>
            <div style="font-size:0.72rem;color:#15803d;">{normal_n:,} pacientes</div>
        </div>""", unsafe_allow_html=True)
    with k5:
        st.markdown(f"""<div style="background:#faf5ff;border:1px solid #d8b4fe;border-radius:14px;
            padding:18px 16px;text-align:center;">
            <div style="font-size:0.68rem;color:#7e22ce;font-family:'DM Mono',monospace;
                        text-transform:uppercase;letter-spacing:1px;">🟡 Sobrepeso/obesidad</div>
            <div style="font-family:'Playfair Display',serif;font-size:2rem;color:#7e22ce;
                        font-weight:700;margin:6px 0;">{pct(exceso_n)}</div>
            <div style="font-size:0.72rem;color:#7e22ce;">{exceso_n:,} casos</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════════════
    # PANEL 2 — TENDENCIA + PANEL 5 — ALERTAS
    # ════════════════════════════════════════════════════════════════════════
    col_tend, col_alertas = st.columns([3,2])

    with col_tend:
        st.markdown('<div class="section-title">Tendencia en el tiempo</div>', unsafe_allow_html=True)
        tend = dff.copy()
        tend["mes"] = tend["fecha"].dt.to_period("M").astype(str)
        tend_g = tend.groupby(["mes","estado_nutricional"]).size().reset_index(name="n")
        tend_tot = tend.groupby("mes").size().reset_index(name="total")
        tend_g = tend_g.merge(tend_tot, on="mes")
        tend_g["pct"] = (tend_g["n"] / tend_g["total"] * 100).round(1)
        estados_tend = ["Desnutrición severa","Desnutrición moderada","Desnutrición leve","Sobrepeso","Obesidad"]
        col_map_t = {
            "Desnutrición severa":"#ef4444","Desnutrición moderada":"#f97316",
            "Desnutrición leve":"#f59e0b","Sobrepeso":"#a855f7","Obesidad":"#4c1d95"
        }
        fig_tend = px.line(
            tend_g[tend_g["estado_nutricional"].isin(estados_tend)],
            x="mes", y="pct", color="estado_nutricional",
            color_discrete_map=col_map_t,
            labels={"pct":"Prevalencia (%)","mes":"Mes","estado_nutricional":"Estado"},
        )
        fig_tend.update_traces(line=dict(width=2.5), mode="lines+markers")
        fig_tend.update_layout(
            height=280, margin=dict(l=0,r=0,t=10,b=0),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(color="#9580c0", gridcolor="#f0e8ff", tickangle=-30),
            yaxis=dict(color="#9580c0", gridcolor="#f0e8ff", title="Prevalencia (%)"),
            legend=dict(font=dict(color="#4a3570",size=9), bgcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig_tend, use_container_width=True)

    with col_alertas:
        st.markdown('<div class="section-title">Panel de alertas</div>', unsafe_allow_html=True)
        ninos_riesgo = len(dff[(dff["grupo_etario"]=="Primera infancia (0-5 años)") &
                               (dff["estado_nutricional"].isin(["Desnutrición severa","Desnutrición moderada"]))])
        perdida_peso = len(dff[dff.get("perdida_peso_reciente", pd.Series([False]*len(dff)))==True])
        gestantes_sin_ctrl = len(dff[(dff.get("gestante", pd.Series([False]*len(dff)))==True) &
                                      (dff.get("control_nutricional", pd.Series([True]*len(dff)))==False)])
        inseg_sev = len(dff[dff.get("inseguridad_alimentaria", pd.Series(["No"]*len(dff))).isin(["Inseguridad severa","Inseguridad moderada"])])
        sin_agua = len(dff[dff.get("agua_potable", pd.Series(["Ok"]*len(dff))).isin(["Sin acceso a agua segura","Agua de río, lluvia o sin tratamiento"])])

        alertas = [
            ("🔴", f"{severos_n} casos de desnutrición severa", "alerta-critica", severos_n > 0),
            ("🔴", f"{ninos_riesgo} niños < 5 años con desnutrición", "alerta-critica", ninos_riesgo > 0),
            ("🔴", f"{perdida_peso} pacientes con pérdida de peso reciente", "alerta-critica", perdida_peso > 0),
            ("🟠", f"{gestantes_sin_ctrl} gestantes sin control nutricional", "alerta-moderada", gestantes_sin_ctrl > 0),
            ("🟠", f"{inseg_sev} pacientes en inseguridad alimentaria", "alerta-moderada", inseg_sev > 0),
            ("🟠", f"{sin_agua} pacientes sin acceso a agua segura", "alerta-moderada", sin_agua > 0),
        ]
        for emoji, msg, css, activa in alertas:
            if activa:
                st.markdown(f'<div class="{css}" style="padding:10px 14px;margin:6px 0;">{emoji} {msg}</div>', unsafe_allow_html=True)
        if not any(a[3] for a in alertas):
            st.markdown('<div class="recom-item">✅ Sin alertas activas en el período seleccionado</div>', unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════════════
    # PANEL 3 — MAPA DE RIESGO NACIONAL
    # ════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="section-title">Mapa de riesgo nutricional — Colombia</div>', unsafe_allow_html=True)

    # Calcular prevalencia por departamento
    prev_dpto = dff.groupby("departamento").apply(
        lambda x: pd.Series({
            "total": len(x),
            "desnutricion": len(x[x["estado_nutricional"].str.contains("Desnutrición")]),
            "obesidad": len(x[x["estado_nutricional"].isin(["Sobrepeso","Obesidad"])]),
        })
    ).reset_index()
    prev_dpto["prev_desnut"] = (prev_dpto["desnutricion"]/prev_dpto["total"]*100).round(1)
    prev_dpto["prev_obesidad"] = (prev_dpto["obesidad"]/prev_dpto["total"]*100).round(1)
    prev_dpto["lat"] = prev_dpto["departamento"].map(lambda d: COORD_DEPTOS.get(d,(4,-74))[0])
    prev_dpto["lon"] = prev_dpto["departamento"].map(lambda d: COORD_DEPTOS.get(d,(4,-74))[1])
    prev_dpto["riesgo"] = pd.cut(prev_dpto["prev_desnut"],
        bins=[-1,10,20,30,100], labels=["Bajo (<10%)","Moderado (10-20%)","Alto (20-30%)","Crítico (>30%)"])

    col_mapa, col_tabla_mapa = st.columns([3,2])
    with col_mapa:
        indicador_mapa = st.radio("Ver en mapa",
            ["Prevalencia desnutrición","Prevalencia sobrepeso/obesidad"], horizontal=True)
        col_ind = "prev_desnut" if "desnutrición" in indicador_mapa else "prev_obesidad"
        titulo_ind = "Desnutrición (%)" if "desnutrición" in indicador_mapa else "Sobrepeso/obesidad (%)"
        escala = [[0,"#f0faf4"],[0.3,"#f59e0b"],[0.6,"#f97316"],[1,"#b91c1c"]] if "desnutrición" in indicador_mapa else [[0,"#f0faf4"],[0.3,"#e9d5ff"],[0.6,"#a855f7"],[1,"#4c1d95"]]

        fig_mapa = px.scatter_geo(
            prev_dpto,
            lat="lat", lon="lon",
            size=col_ind,
            color=col_ind,
            hover_name="departamento",
            hover_data={"lat":False,"lon":False,"total":True,"prev_desnut":True,"prev_obesidad":True},
            color_continuous_scale=escala,
            size_max=40,
            labels={col_ind: titulo_ind, "total":"Total evaluados","prev_desnut":"Desnut. (%)","prev_obesidad":"Obesidad (%)"},
        )
        fig_mapa.update_geos(
            scope="south america",
            center={"lat":4.5,"lon":-74.3},
            projection_scale=3.5,
            showland=True, landcolor="#f7f4fb",
            showocean=True, oceancolor="#e8f4fd",
            showcoastlines=True, coastlinecolor="#c4b5fd",
            showcountries=True, countrycolor="#c4b5fd",
        )
        fig_mapa.update_layout(
            height=420, margin=dict(l=0,r=0,t=10,b=0),
            paper_bgcolor="rgba(0,0,0,0)",
            coloraxis_colorbar=dict(
                title=titulo_ind,
            ),
        )
        st.plotly_chart(fig_mapa, use_container_width=True)

    with col_tabla_mapa:
        st.caption("Prevalencia por departamento — ordenado por riesgo")
        tabla_mapa = prev_dpto[["departamento","total","prev_desnut","prev_obesidad","riesgo"]].copy()
        tabla_mapa.columns = ["Departamento","Evaluados","Desnut. (%)","Obesidad (%)","Nivel riesgo"]
        tabla_mapa = tabla_mapa.sort_values("Desnut. (%)", ascending=False)
        st.dataframe(tabla_mapa, use_container_width=True, height=380, hide_index=True)

    # ════════════════════════════════════════════════════════════════════════
    # PANEL 4 — DISTRIBUCIÓN POR GRUPO ETARIO
    # ════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="section-title">Distribución por grupo etario — Normativa Colombia</div>', unsafe_allow_html=True)

    col_bar, col_pie_g = st.columns(2)
    with col_bar:
        grupo_est = dff.groupby(["grupo_etario","estado_nutricional"]).size().reset_index(name="n")
        grupo_tot = dff.groupby("grupo_etario").size().reset_index(name="total")
        grupo_est = grupo_est.merge(grupo_tot, on="grupo_etario")
        grupo_est["pct"] = (grupo_est["n"]/grupo_est["total"]*100).round(1)
        col_estados = {
            "Desnutrición severa":"#ef4444","Desnutrición moderada":"#f97316",
            "Desnutrición leve":"#f59e0b","Normal":"#22c55e",
            "Sobrepeso":"#a855f7","Obesidad":"#4c1d95"
        }
        fig_g = px.bar(grupo_est, x="grupo_etario", y="pct", color="estado_nutricional",
            color_discrete_map=col_estados,
            labels={"pct":"Prevalencia (%)","grupo_etario":"Grupo etario","estado_nutricional":"Estado"},
            barmode="stack")
        fig_g.update_layout(
            height=320, margin=dict(l=0,r=0,t=10,b=0),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(color="#9580c0", tickangle=-20, tickfont=dict(size=9)),
            yaxis=dict(color="#9580c0", gridcolor="#f0e8ff", title="%"),
            legend=dict(font=dict(color="#4a3570",size=9), bgcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig_g, use_container_width=True)

    with col_pie_g:
        st.caption("Resumen crítico por grupo etario")
        for grupo in list(GRUPOS_ETARIOS.keys()):
            sub = dff[dff["grupo_etario"]==grupo]
            if len(sub) == 0: continue
            desnut_g = len(sub[sub["estado_nutricional"].str.contains("Desnutrición")])
            exceso_g = len(sub[sub["estado_nutricional"].isin(["Sobrepeso","Obesidad"])])
            pct_d = desnut_g/len(sub)*100
            pct_e = exceso_g/len(sub)*100
            color_ind = "#ef4444" if pct_d > 25 else "#f59e0b" if pct_d > 15 else "#22c55e"
            st.markdown(f"""
            <div class="recom-item" style="padding:10px 14px;border-left:3px solid {color_ind};">
                <strong style="font-size:0.8rem;">{grupo}</strong>
                <div style="font-size:0.75rem;color:#6b3fbf;margin-top:4px;">
                    {len(sub)} evaluados · Desnut.: <strong>{pct_d:.1f}%</strong> · Exceso peso: <strong>{pct_e:.1f}%</strong>
                </div>
            </div>""", unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════════════
    # PANEL 6 — DETERMINANTES SOCIALES
    # ════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="section-title">Determinantes sociales y estado nutricional</div>', unsafe_allow_html=True)

    col_d1, col_d2 = st.columns(2)
    with col_d1:
        # Acceso al agua vs desnutrición
        agua_desnut = dff.groupby("agua_potable").apply(
            lambda x: pd.Series({
                "total": len(x),
                "desnutricion_pct": len(x[x["estado_nutricional"].str.contains("Desnutrición")])/len(x)*100 if len(x)>0 else 0
            })
        ).reset_index()
        agua_desnut = agua_desnut.sort_values("desnutricion_pct", ascending=True)
        fig_agua = go.Figure(go.Bar(
            y=agua_desnut["agua_potable"],
            x=agua_desnut["desnutricion_pct"].round(1),
            orientation="h",
            marker=dict(color=agua_desnut["desnutricion_pct"],
                colorscale=[[0,"#f0faf4"],[0.5,"#f59e0b"],[1,"#ef4444"]]),
            text=agua_desnut["desnutricion_pct"].round(1).astype(str)+"%",
            textposition="outside",
        ))
        fig_agua.update_layout(
            title=dict(text="Acceso al agua vs desnutrición",font=dict(size=11,color="#9580c0")),
            height=260, margin=dict(l=0,r=0,t=30,b=0),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(color="#9580c0", gridcolor="#f0e8ff", title="Desnutrición (%)"),
            yaxis=dict(color="#4a3570", tickfont=dict(size=9)),
            coloraxis_showscale=False,
        )
        st.plotly_chart(fig_agua, use_container_width=True)

    with col_d2:
        # Inseguridad alimentaria vs desnutrición
        inseg_desnut = dff.groupby("inseguridad_alimentaria").apply(
            lambda x: pd.Series({
                "total": len(x),
                "desnutricion_pct": len(x[x["estado_nutricional"].str.contains("Desnutrición")])/len(x)*100 if len(x)>0 else 0
            })
        ).reset_index()
        inseg_desnut = inseg_desnut.sort_values("desnutricion_pct", ascending=True)
        fig_inseg = go.Figure(go.Bar(
            y=inseg_desnut["inseguridad_alimentaria"],
            x=inseg_desnut["desnutricion_pct"].round(1),
            orientation="h",
            marker=dict(color=inseg_desnut["desnutricion_pct"],
                colorscale=[[0,"#f0faf4"],[0.5,"#f59e0b"],[1,"#ef4444"]]),
            text=inseg_desnut["desnutricion_pct"].round(1).astype(str)+"%",
            textposition="outside",
        ))
        fig_inseg.update_layout(
            title=dict(text="Inseguridad alimentaria vs desnutrición",font=dict(size=11,color="#9580c0")),
            height=260, margin=dict(l=0,r=0,t=30,b=0),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(color="#9580c0", gridcolor="#f0e8ff", title="Desnutrición (%)"),
            yaxis=dict(color="#4a3570", tickfont=dict(size=9)),
            coloraxis_showscale=False,
        )
        st.plotly_chart(fig_inseg, use_container_width=True)

    # ════════════════════════════════════════════════════════════════════════
    # PANEL 7 — PANEL PREDICTIVO
    # ════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="section-title">Panel predictivo — proyección próximos 3 meses</div>', unsafe_allow_html=True)
    st.markdown('<div class="recom-item"><strong>Metodología:</strong> Proyección basada en tendencia lineal de los últimos 6 meses. Los valores son estimaciones estadísticas — no reemplazan el análisis clínico y epidemiológico especializado.</div>', unsafe_allow_html=True)

    # Calcular tendencia real de los datos
    tend_mens = dff.copy()
    tend_mens["mes_num"] = tend_mens["fecha"].dt.to_period("M").apply(lambda x: x.ordinal)
    tend_mens["es_desnut"] = tend_mens["estado_nutricional"].str.contains("Desnutrición").astype(int)
    tend_mens["es_exceso"] = tend_mens["estado_nutricional"].isin(["Sobrepeso","Obesidad"]).astype(int)

    # Tendencia por mes
    t_group = tend_mens.groupby("mes_num").agg(
        total=("es_desnut","count"),
        desnut=("es_desnut","sum"),
        exceso=("es_exceso","sum"),
    ).reset_index()
    t_group["pct_desnut"] = t_group["desnut"]/t_group["total"]*100
    t_group["pct_exceso"] = t_group["exceso"]/t_group["total"]*100

    # Proyección simple (tendencia lineal)
    if len(t_group) >= 2:
        x = t_group["mes_num"].values
        y_d = t_group["pct_desnut"].values
        y_e = t_group["pct_exceso"].values
        coef_d = np.polyfit(x, y_d, 1)
        coef_e = np.polyfit(x, y_e, 1)
        x_fut = np.array([x[-1]+1, x[-1]+2, x[-1]+3])
        proy_d = np.clip(np.polyval(coef_d, x_fut), 0, 100)
        proy_e = np.clip(np.polyval(coef_e, x_fut), 0, 100)
        tend_d = "↑ Aumentando" if coef_d[0] > 0 else "↓ Disminuyendo"
        tend_e = "↑ Aumentando" if coef_e[0] > 0 else "↓ Disminuyendo"
    else:
        proy_d = np.array([desnut_n/total*100]*3)
        proy_e = np.array([exceso_n/total*100]*3)
        tend_d = "→ Estable"
        tend_e = "→ Estable"

    col_p1, col_p2, col_p3, col_p4 = st.columns(4)
    color_tend_d = "#ef4444" if "Aumentando" in tend_d else "#22c55e"
    color_tend_e = "#a855f7" if "Aumentando" in tend_e else "#22c55e"

    with col_p1:
        st.markdown(f"""<div class="result-card result-{'malo' if 'Aumentando' in tend_d else 'normal'}">
            <h3>Desnutrición — mes 1</h3>
            <p>Estimado: <strong>{proy_d[0]:.1f}%</strong></p>
        </div>""", unsafe_allow_html=True)
    with col_p2:
        st.markdown(f"""<div class="result-card result-{'malo' if 'Aumentando' in tend_d else 'normal'}">
            <h3>Desnutrición — mes 2</h3>
            <p>Estimado: <strong>{proy_d[1]:.1f}%</strong></p>
        </div>""", unsafe_allow_html=True)
    with col_p3:
        st.markdown(f"""<div class="result-card result-{'malo' if 'Aumentando' in tend_d else 'normal'}">
            <h3>Desnutrición — mes 3</h3>
            <p>Estimado: <strong>{proy_d[2]:.1f}%</strong></p>
        </div>""", unsafe_allow_html=True)
    with col_p4:
        st.markdown(f"""<div class="result-card result-sobrepeso">
            <h3>Exceso peso — mes 3</h3>
            <p>Estimado: <strong>{proy_e[2]:.1f}%</strong></p>
        </div>""", unsafe_allow_html=True)

    # Gráfica proyección
    meses_hist = [str(p) for p in pd.period_range(tend_mens["fecha"].min().to_period("M"),
                                                    tend_mens["fecha"].max().to_period("M"), freq="M")]
    ult_mes = pd.Period(meses_hist[-1], freq="M") if meses_hist else pd.Period("2024-12","M")
    meses_proy = [str(ult_mes+i) for i in range(1,4)]

    fig_proy = go.Figure()
    if len(t_group) > 0:
        fig_proy.add_trace(go.Scatter(
            x=meses_hist[:len(t_group)], y=t_group["pct_desnut"].round(1),
            name="Desnutrición histórica", line=dict(color="#ef4444",width=2.5), mode="lines+markers"))
        fig_proy.add_trace(go.Scatter(
            x=meses_hist[:len(t_group)], y=t_group["pct_exceso"].round(1),
            name="Exceso peso histórico", line=dict(color="#a855f7",width=2.5), mode="lines+markers"))
    fig_proy.add_trace(go.Scatter(
        x=meses_proy, y=proy_d.round(1),
        name="Proyección desnutrición",
        line=dict(color="#ef4444",width=2,dash="dash"), mode="lines+markers",
        marker=dict(symbol="diamond",size=8)))
    fig_proy.add_trace(go.Scatter(
        x=meses_proy, y=proy_e.round(1),
        name="Proyección exceso peso",
        line=dict(color="#a855f7",width=2,dash="dash"), mode="lines+markers",
        marker=dict(symbol="diamond",size=8)))
    fig_proy.add_vrect(
        x0=meses_proy[0], x1=meses_proy[-1],
        fillcolor="rgba(168,85,247,0.06)", line_width=0,
        annotation_text="Proyección", annotation_position="top left",
        annotation_font=dict(size=10,color="#9580c0"))
    fig_proy.update_layout(
        height=280, margin=dict(l=0,r=0,t=20,b=0),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(color="#9580c0", gridcolor="#f0e8ff", tickangle=-30),
        yaxis=dict(color="#9580c0", gridcolor="#f0e8ff", title="Prevalencia (%)"),
        legend=dict(font=dict(color="#4a3570",size=9), bgcolor="rgba(0,0,0,0)"),
    )
    st.plotly_chart(fig_proy, use_container_width=True)

    # Alertas predictivas
    if proy_d[2] > desnut_n/total*100 + 3:
        st.markdown(f'<div class="alerta-critica">⚠ ALERTA PREDICTIVA — La prevalencia de desnutrición podría aumentar a <strong>{proy_d[2]:.1f}%</strong> en los próximos 3 meses. Se recomienda fortalecer intervenciones preventivas.</div>', unsafe_allow_html=True)
    if proy_e[2] > exceso_n/total*100 + 3:
        st.markdown(f'<div class="alerta-moderada">⚠ ALERTA PREDICTIVA — El exceso de peso podría aumentar a <strong>{proy_e[2]:.1f}%</strong>. Reforzar programas de actividad física y orientación alimentaria.</div>', unsafe_allow_html=True)

    # Departamentos críticos
    st.markdown('<div class="section-title">Departamentos en mayor riesgo</div>', unsafe_allow_html=True)
    top_riesgo = prev_dpto.nlargest(5, "prev_desnut")[["departamento","total","prev_desnut","prev_obesidad"]]
    top_riesgo.columns = ["Departamento","Evaluados","Desnutrición (%)","Obesidad (%)"]
    c1,c2 = st.columns(2)
    with c1:
        st.caption("Top 5 mayor desnutrición")
        st.dataframe(top_riesgo, use_container_width=True, hide_index=True)
    with c2:
        top_obesidad = prev_dpto.nlargest(5, "prev_obesidad")[["departamento","total","prev_desnut","prev_obesidad"]]
        top_obesidad.columns = ["Departamento","Evaluados","Desnutrición (%)","Obesidad (%)"]
        st.caption("Top 5 mayor obesidad")
        st.dataframe(top_obesidad, use_container_width=True, hide_index=True)

elif modulo == "Registro Masivo":
    st.markdown('<div class="section-title">Plantillas por grupo etario — Normativa Colombia</div>', unsafe_allow_html=True)
    st.markdown('<div class="recom-item"><strong>Selecciona el grupo etario</strong> para descargar la plantilla con los campos específicos y el formato adecuado según la normativa colombiana vigente.</div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="recom-item">
        <strong>Grupos etarios según Normativa Colombia vigente:</strong><br>
        <span style="font-size:0.78rem;color:#5a8a6a;">
        Primera infancia (0-5 años) · Ley 1098/2006 &nbsp;|&nbsp;
        Infancia (6-11 años) · Ley 1098/2006 &nbsp;|&nbsp;
        Adolescencia (12-17 años) · Ley 1098/2006 &nbsp;|&nbsp;
        Juventud (18-28 años) · Ley 1622/2013 &nbsp;|&nbsp;
        Adultez (29-59 años) · Ley 1751/2015 &nbsp;|&nbsp;
        Persona mayor (≥60 años) · Ley 1251/2008
        </span>
    </div>
    """, unsafe_allow_html=True)
    grupo_plantilla = st.selectbox("Grupo etario para la plantilla", list(GRUPOS_ETARIOS.keys()))

    def crear_plantilla_grupo(grupo):
        base = {
            "nombre":["Ejemplo Paciente"],"fecha_evaluacion":["2026-03-19"],
            "grupo_etario":[grupo],"departamento":["Chocó"],"municipio":["Quibdó"],
            "sexo":["Femenino"],"edad_anos":[4],"peso_kg":[14.5],"talla_cm":[98.0],
            "perimetro_braquial_cm":[13.5],"perimetro_cintura_cm":[50.0],"perimetro_cadera_cm":[54.0],
            "estrato":["1"],"regimen_salud":["Subsidiado (SISBEN)"],
            "tipo_vivienda":["Casa propia"],"personas_hogar":[5],
            "escolaridad_cuidador":["Primaria completa"],"etnia":["Afrocolombiano"],
            "comidas_dia":[3],
            "disponibilidad_alimentos":["A veces"],"fuente_alimentos":["Compra en tienda o supermercado"],
            "acceso_economico":["A veces no alcanza"],"acceso_fisico":["15-30 minutos"],
            "programas_alimentarios":["PAE (escolar)"],
            "decision_alimentaria":["La madre"],"inseguridad_ELCSA":["Inseguridad moderada"],
            "agua_potable":["Agua de río, lluvia o sin tratamiento"],
            "almacenamiento_alimentos":["Inadecuado (sin refrigeración)"],
            "preparacion_alimentos":["Fogón de leña o carbón"],
            "condicion_absorcion":["Parasitosis intestinal"],"desparasitacion":["Hace más de 1 año"],
            "saneamiento_basico":["Letrina"],"lavado_manos":["A veces"],
            "enfermedades_personales":["Anemia"],"antecedentes_familiares":["Desnutrición infantil"],
            "hospitalizacion_previa":["No"],"medicamentos":["Sulfato ferroso"],
            "alergias_intolerancias":["Ninguna conocida"],
        }
        for grupo_alim in GRUPOS_ALIMENTOS.keys():
            base[f"consumo_{grupo_alim[:20].replace(' ','_')}"] = ["1-2 veces/semana"]
        base["observaciones"] = [""]
        return pd.DataFrame(base)

    df_plantilla = crear_plantilla_grupo(grupo_plantilla)

    # Formato Excel con colores
    buf = io.BytesIO()
    if OPENPYXL_OK:
        try:
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                df_plantilla.to_excel(writer, index=False, sheet_name=grupo_plantilla[:28])
                ws = writer.sheets[grupo_plantilla[:28]]
                sc = {"identificacion":"D8C5F5","antropometria":"C5E8F5","sociodemografico":"C5F5D8","seguridad_alimentaria":"F5E8C5","antecedentes":"F5C5C5","consumo":"F5F5C5"}
                hf = PatternFill(start_color="2D1B69",end_color="2D1B69",fill_type="solid")
                hfont = Font(bold=True,color="FFFFFF",size=10)
                bd = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
                for ci,cn in enumerate(df_plantilla.columns,1):
                    c=ws.cell(row=1,column=ci); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=bd
                    fc=sc["identificacion"] if ci<=8 else sc["antropometria"] if ci<=14 else sc["sociodemografico"] if ci<=20 else sc["seguridad_alimentaria"] if ci<=30 else sc["antecedentes"] if ci<=38 else sc["consumo"]
                    dc=ws.cell(row=2,column=ci); dc.fill=PatternFill(start_color=fc,end_color=fc,fill_type="solid"); dc.border=bd; dc.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
                    ws.column_dimensions[get_column_letter(ci)].width=max(15,min(35,len(cn)+5))
                ws.row_dimensions[1].height=35; ws.row_dimensions[2].height=20; ws.freeze_panes="A2"
        except Exception:
            buf = io.BytesIO()
            df_plantilla.to_excel(buf, index=False)
    else:
        df_plantilla.to_excel(buf, index=False)


    st.download_button(
        f"Descargar plantilla — {grupo_plantilla}",
        data=buf.getvalue(),
        file_name=f"NutriVida_plantilla_{grupo_plantilla[:20].replace(' ','_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.markdown('<div class="section-title">Leyenda de colores de la plantilla</div>', unsafe_allow_html=True)
    leyenda = [
        ("Identificación y datos básicos","#D8C5F5"),
        ("Datos antropométricos","#C5E8F5"),
        ("Datos sociodemográficos","#C5F5D8"),
        ("Seguridad alimentaria (5 componentes)","#F5E8C5"),
        ("Antecedentes personales y familiares","#F5C5C5"),
        ("Frecuencia de consumo por grupos","#F5F5C5"),
    ]
    c1,c2,c3 = st.columns(3)
    for i,(nombre_sec,color) in enumerate(leyenda):
        with [c1,c2,c3][i%3]:
            st.markdown(f'<div style="background:#{color};border:1px solid #ccc;border-radius:8px;padding:8px 12px;margin:4px 0;font-size:0.78rem;color:#1a1025;"><strong>{nombre_sec}</strong></div>', unsafe_allow_html=True)

    st.markdown('<div class="section-title">Cargar archivo diligenciado</div>', unsafe_allow_html=True)
    archivo = st.file_uploader("Subir archivo Excel o CSV con datos", type=["xlsx","csv"])
    if archivo:
        try:
            datos = pd.read_csv(archivo) if archivo.name.endswith(".csv") else pd.read_excel(archivo)
            st.success(f"{len(datos)} registros cargados correctamente")
            datos["imc"] = datos.apply(lambda r: calcular_imc(r["peso_kg"], r["talla_cm"]), axis=1)
            datos["clasificacion"] = datos.apply(lambda r: clasificar_nino(r["imc"], r.get("edad_anos",4)*12)[0], axis=1)
            datos["alerta"] = datos["clasificacion"].apply(lambda c: "URGENTE" if "severa" in c.lower() else ("PRIORITARIO" if "moderada" in c.lower() else "RUTINA"))
            k1,k2,k3 = st.columns(3)
            k1.metric("Total procesados", len(datos))
            k2.metric("Casos urgentes", len(datos[datos["alerta"]=="URGENTE"]))
            k3.metric("Casos prioritarios", len(datos[datos["alerta"]=="PRIORITARIO"]))
            st.dataframe(datos[["nombre","municipio","departamento","edad_anos","peso_kg","imc","clasificacion","alerta"]], use_container_width=True)
            out = io.BytesIO(); datos.to_excel(out, index=False)
            st.download_button("Exportar resultados procesados", data=out.getvalue(), file_name="NutriVida_resultados.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Error procesando archivo: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 5 — ACERCA DEL SISTEMA
# ═══════════════════════════════════════════════════════════════════════════════
elif modulo == "Acerca del sistema":
    c1,c2 = st.columns([3,2])
    with c1:
        st.markdown('<div class="section-title">NutriVida Colombia v7.0 — descripción</div>', unsafe_allow_html=True)
        for t,d,url in [
            ("Propósito","Sistema integral de vigilancia y evaluación nutricional para profesionales de salud del sector público colombiano. Evaluación por grupos etarios según normativa vigente.",""),
            ("Marco normativo","Ley 1098/2006, Ley 1622/2013, Ley 1251/2008, Ley 1751/2015, OMS/OPS 2006, ENSIN 2015, Res. 2465/2016, ELCSA, Protocolo AIEPI.","https://www.minsalud.gov.co"),
            ("Seguridad alimentaria","5 componentes: disponibilidad, acceso, consumo, calidad e inocuidad, aprovechamiento biológico. CONPES 113/2007.","https://www.fao.org/colombia"),
            ("ENSIN 2015","Encuesta Nacional de la Situación Nutricional. ICBF, MinSalud, INS, Profamilia.","https://www.icbf.gov.co/bienestar/nutricion/encuesta-nacional-situacion-nutricional"),
            ("SIVIGILA","Sistema Nacional de Vigilancia en Salud Pública. INS Colombia.","https://www.ins.gov.co/Noticias/Paginas/sivigila.aspx"),
            ("Tecnología","Python + Streamlit + Pandas + Plotly + OpenPyXL. Código abierto, sin licencias comerciales.",""),
        ]:
            link = f' <a href="{url}" target="_blank" style="color:#a855f7;font-size:0.72rem;">Ver fuente →</a>' if url else ""
            st.markdown(f'<div class="recom-item"><strong>{t}</strong> — {d}{link}</div>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="section-title">Autora y derechos</div>', unsafe_allow_html=True)
        st.markdown("""<div class="recom-item" style="background:linear-gradient(135deg,#faf5ff,#f3e8ff);border-color:#d8b4fe;text-align:center;padding:28px;">
        <div style="font-family:'Playfair Display',serif;font-size:1.2rem;color:#2d1b69;font-weight:700;margin-bottom:6px;">Maira Alejandra Carrillo Florez</div>
        <div style="font-size:0.85rem;color:#6b3fbf;margin-bottom:2px;">Nutricionista - Dietista</div>
        <div style="font-size:0.78rem;color:#9580c0;">Universidad de Pamplona · 2026</div>
        <hr style="border-color:#e9d5ff;margin:16px 0;">
        <div style="font-size:0.72rem;color:#7c3aed;font-family:'DM Mono',monospace;line-height:2;">
        © 2026 · Todos los derechos reservados<br>Ley 23 de 1982 · Decisión Andina 351<br>Registro DNDA — En trámite 2026
        </div></div>""", unsafe_allow_html=True)

        st.markdown('<div class="section-title">Referencias con enlace</div>', unsafe_allow_html=True)
        for c,d,url in [
            ("ENSIN 2015","Situación Nutricional Colombia","https://www.icbf.gov.co/bienestar/nutricion/encuesta-nacional-situacion-nutricional"),
            ("OMS/OPS 2006","Curvas crecimiento infantil","https://www.who.int/tools/child-growth-standards"),
            ("Res. 2465/2016","Indicadores antropométricos","https://www.minsalud.gov.co/Normatividad_Nuevo/Resoluci%C3%B3n%202465%20de%202016.pdf"),
            ("CONPES 113/2007","Política SAN Colombia","https://www.dnp.gov.co/CONPES/documentos-conpes/Paginas/documentos-conpes.aspx"),
            ("SIVIGILA","Vigilancia Salud Pública","https://www.ins.gov.co/Noticias/Paginas/sivigila.aspx"),
            ("SISVAN","Vigilancia Alimentaria","https://www.minsalud.gov.co/salud/publica/PENT/Paginas/sistema-de-vigilancia-alimentaria-y-nutricional.aspx"),
        ]:
            st.markdown(f'<div class="recom-item" style="padding:8px 14px;"><strong>{c}</strong> — {d} <a href="{url}" target="_blank" style="color:#a855f7;font-size:0.72rem;">Ver →</a></div>', unsafe_allow_html=True)

st.markdown("""<div class="footer-gov">
NutriVida Colombia v7.0 · Sistema Integral de Evaluación y Seguimiento Nutricional · República de Colombia<br>
Ministerio de Salud y Protección Social · ICBF · Datos con fines demostrativos<br><br>
© 2026 Maira Alejandra Carrillo Florez — Todos los derechos reservados<br>
Nutricionista - Dietista · Universidad de Pamplona<br>
Protegido bajo la Ley 23 de 1982 · Decisión Andina 351 · Registro DNDA en trámite
</div>""", unsafe_allow_html=True)
